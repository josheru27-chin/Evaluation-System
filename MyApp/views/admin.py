from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from io import TextIOWrapper
from datetime import datetime
import csv
from collections import defaultdict
from django.db.models import Prefetch, Case, When, IntegerField
from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.core.signing import TimestampSigner
from django.template.loader import render_to_string
from django.urls import reverse
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.text import slugify
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
import re
from difflib import SequenceMatcher

from ..admin_roles import get_admin_role, admin_landing_page, can_access_role
from django.contrib.auth.models import User, Group
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..models import (
    Department,
    FacultyMember,
    DepartmentHead,
    EvaluationOfficer,
    EvaluationSchedule,
    FacultyEvaluation,
    FacultyEvaluationResponse,
    OfficeEvaluation,
    OfficeEvaluationResponse,
    HeadEvaluation,
    HeadEvaluationResponse,
    SEFSETUploadBatch,
    SEFSETMatchedResult,
)

LOGIN_LINK_MAX_AGE = 300
LINK_SALT = "faculty-eval-login"


def _get_open_schedule():
    now = timezone.localtime(timezone.now())
    return (
        EvaluationSchedule.objects
        .filter(start_datetime__lte=now, end_datetime__gte=now)
        .order_by("start_datetime")
        .first()
    )


DEPARTMENT_MAP = {
    "UITC": "University Information Technology Center",
    "DED": "Department of Industrial Education",
    "DIT": "Department of Industrial Technology",
    "DLA": "Department of Liberal Arts",
    "DOE": "Department of Engineering",
    "DMS": "Department of Math and Science",
}

def _clean_import_header(value):
    return re.sub(r"[^A-Z0-9]", "", str(value or "").strip().upper())


def _clean_import_value(value):
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _get_header_index(headers, *possible_names):
    normalized_headers = [_clean_import_header(header) for header in headers]
    possible_names = [_clean_import_header(name) for name in possible_names]

    for possible_name in possible_names:
        if possible_name in normalized_headers:
            return normalized_headers.index(possible_name)

    return None


def _get_row_value(row, index):
    if index is None:
        return ""

    if index >= len(row):
        return ""

    return _clean_import_value(row[index])

def _admin_context(active_page, extra=None):
    context = {"active_page": active_page}
    if extra:
        context.update(extra)
    return context


def _ordered_response_queryset(model):
    return model.objects.annotate(
        section_order=Case(
            When(section_code="management_teaching_learning", then=0),
            When(section_code="content_knowledge_pedagogy_technology", then=1),
            When(section_code="commitment_transparency", then=2),
            default=99,
            output_field=IntegerField(),
        )
    ).order_by("section_order", "question_number")

def _replace_faculty_from_department_sheet(ws, department, schedule):
    """
    Supports department sheets like:
    NAME | GSFE EMAIL | RANK | COURSE | PROGRAM_YEAR

    Also supports:
    ID NUMBER | NAME | EMAIL
    FIRST NAME | LAST NAME | EMAIL
    """

    FacultyMember.objects.filter(schedule=schedule, department=department).delete()

    headers = [
        cell.value for cell in ws[1]
    ]

    name_index = _get_header_index(headers, "NAME", "FULL NAME", "FACULTY NAME")
    first_name_index = _get_header_index(headers, "FIRST NAME", "FIRSTNAME")
    last_name_index = _get_header_index(headers, "LAST NAME", "LASTNAME")
    id_index = _get_header_index(headers, "ID NUMBER", "ID NO", "ID", "EMPLOYEE ID")
    email_index = _get_header_index(headers, "GSFE EMAIL", "EMAIL", "EMAIL ADDRESS")

    rank_index = _get_header_index(
        headers,
        "RANK",
        "FACULTY RANK",
        "ACADEMIC RANK",
        "CURRENT FACULTY RANK",
        "POSITION",
    )

    course_index = _get_header_index(
        headers,
        "COURSE",
        "COURSE CODE",
        "COURSE CODE TITLE",
        "COURSE CODE/TITLE",
        "COURSE TITLE",
    )

    program_year_index = _get_header_index(
        headers,
        "PROGRAM_YEAR",
        "PROGRAM YEAR",
        "YEAR",
        "YEAR LEVEL",
    )

    faculty_to_create = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        id_number = _get_row_value(row, id_index)
        name = _get_row_value(row, name_index)
        email = _get_row_value(row, email_index)
        academic_rank = _get_row_value(row, rank_index)
        course = _get_row_value(row, course_index)
        program_year = _get_row_value(row, program_year_index)

        if not name:
            first_name = _get_row_value(row, first_name_index)
            last_name = _get_row_value(row, last_name_index)
            name = f"{first_name} {last_name}".strip()

        if not name:
            continue

        faculty_to_create.append(
            FacultyMember(
                schedule=schedule,
                department=department,
                id_number=id_number,
                name=name,
                email=email,
                academic_rank=academic_rank,
                course=course,
                program_year=program_year,
            )
        )

    FacultyMember.objects.bulk_create(faculty_to_create)
    return len(faculty_to_create)

def _replace_faculty_from_uploaded_file(uploaded_file, department, schedule):
    """
    Supports:
    - .xlsx
    - .csv

    Accepted columns:
    NAME, GSFE EMAIL, RANK, COURSE, PROGRAM_YEAR
    """

    file_name = uploaded_file.name.lower()

    FacultyMember.objects.filter(schedule=schedule, department=department).delete()
    created_count = 0

    if file_name.endswith(".xlsx"):
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        headers = [
            cell.value for cell in ws[1]
        ]

        name_index = _get_header_index(headers, "NAME", "FULL NAME", "FACULTY NAME")
        first_name_index = _get_header_index(headers, "FIRST NAME", "FIRSTNAME")
        last_name_index = _get_header_index(headers, "LAST NAME", "LASTNAME")
        id_index = _get_header_index(headers, "ID NUMBER", "ID NO", "ID", "EMPLOYEE ID")
        email_index = _get_header_index(headers, "GSFE EMAIL", "EMAIL", "EMAIL ADDRESS")

        rank_index = _get_header_index(
            headers,
            "RANK",
            "FACULTY RANK",
            "ACADEMIC RANK",
            "CURRENT FACULTY RANK",
            "POSITION",
        )

        course_index = _get_header_index(
            headers,
            "COURSE",
            "COURSE CODE",
            "COURSE CODE TITLE",
            "COURSE CODE/TITLE",
            "COURSE TITLE",
        )

        program_year_index = _get_header_index(
            headers,
            "PROGRAM_YEAR",
            "PROGRAM YEAR",
            "YEAR",
            "YEAR LEVEL",
        )

        if name_index is None and (first_name_index is None or last_name_index is None):
            return 0

        faculty_to_create = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = _get_row_value(row, name_index)

            if not name:
                first_name = _get_row_value(row, first_name_index)
                last_name = _get_row_value(row, last_name_index)
                name = f"{first_name} {last_name}".strip()

            if not name:
                continue

            faculty_to_create.append(
                FacultyMember(
                    schedule=schedule,
                    department=department,
                    id_number=_get_row_value(row, id_index),
                    name=name,
                    email=_get_row_value(row, email_index),
                    academic_rank=_get_row_value(row, rank_index),
                    course=_get_row_value(row, course_index),
                    program_year=_get_row_value(row, program_year_index),
                )
            )

        FacultyMember.objects.bulk_create(faculty_to_create)
        created_count = len(faculty_to_create)

    elif file_name.endswith(".csv"):
        decoded_file = TextIOWrapper(uploaded_file.file, encoding="utf-8")
        reader = csv.DictReader(decoded_file)

        field_map = {
            _clean_import_header(field): field
            for field in (reader.fieldnames or [])
        }

        def get_csv_value(row, *names):
            for name in names:
                key = field_map.get(_clean_import_header(name))
                if key:
                    return _clean_import_value(row.get(key))
            return ""

        faculty_to_create = []

        for row in reader:
            name = get_csv_value(row, "NAME", "FULL NAME", "FACULTY NAME")

            if not name:
                first_name = get_csv_value(row, "FIRST NAME", "FIRSTNAME")
                last_name = get_csv_value(row, "LAST NAME", "LASTNAME")
                name = f"{first_name} {last_name}".strip()

            if not name:
                continue

            faculty_to_create.append(
                FacultyMember(
                    schedule=schedule,
                    department=department,
                    id_number=get_csv_value(row, "ID NUMBER", "ID NO", "ID", "EMPLOYEE ID"),
                    name=name,
                    email=get_csv_value(row, "GSFE EMAIL", "EMAIL", "EMAIL ADDRESS"),
                    academic_rank=get_csv_value(row, "RANK", "FACULTY RANK", "ACADEMIC RANK", "POSITION"),
                    course=get_csv_value(row, "COURSE", "COURSE CODE", "COURSE CODE/TITLE", "COURSE TITLE"),
                    program_year=get_csv_value(row, "PROGRAM_YEAR", "PROGRAM YEAR", "YEAR", "YEAR LEVEL"),
                )
            )

        FacultyMember.objects.bulk_create(faculty_to_create)
        created_count = len(faculty_to_create)

    return created_count

def _parse_datetime_local(value):
    if not value:
        return None
    naive_dt = datetime.strptime(value, "%Y-%m-%dT%H:%M")
    return timezone.make_aware(naive_dt, timezone.get_current_timezone())


def _get_latest_schedule_with_uploaded_data():
    schedules = (
        EvaluationSchedule.objects
        .order_by("-start_datetime", "-created_at")
    )

    for schedule in schedules:
        has_faculty = FacultyMember.objects.filter(schedule=schedule).exists()
        has_heads = DepartmentHead.objects.filter(schedule=schedule).exists()

        if has_faculty or has_heads:
            return schedule

    return None


def _get_latest_schedule_with_submitted_evaluations():
    schedules = EvaluationSchedule.objects.order_by("-start_datetime", "-created_at")

    for schedule in schedules:
        has_faculty_eval = FacultyEvaluation.objects.filter(
            schedule=schedule,
            status="submitted"
        ).exists()

        has_office_eval = OfficeEvaluation.objects.filter(
            schedule=schedule,
            status="submitted"
        ).exists()

        has_head_eval = HeadEvaluation.objects.filter(
            schedule=schedule,
            status="submitted"
        ).exists()

        if has_faculty_eval or has_office_eval or has_head_eval:
            return schedule

    return None

from functools import wraps


def role_required(*allowed_roles):
    def decorator(view_func):
        @wraps(view_func)
        @login_required(login_url="admin_login")
        def wrapper(request, *args, **kwargs):
            user = request.user

            if not user.is_staff:
                messages.error(request, "You do not have admin access.")
                return redirect("admin_login")

            if not can_access_role(user, allowed_roles):
                messages.error(request, "You are not allowed to access that panel.")
                return redirect(admin_landing_page(user))

            return view_func(request, *args, **kwargs)

        return wrapper
    return decorator


def admin_required(view_func):
    return role_required("UITC", "ADAA", "STAFF")(view_func)


@role_required("UITC", "ADAA")
def admin_department(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        schedule_id = request.POST.get("schedule_id")
        if not schedule_id:
            messages.error(request, "Please select an evaluation schedule.")
            return redirect("admin_department")

        selected_schedule = get_object_or_404(EvaluationSchedule, id=schedule_id)
        excel_file = request.FILES["excel_file"]

        try:
            wb = load_workbook(excel_file, data_only=True)
        except Exception:
            messages.error(request, "Invalid Excel file. Please upload a valid .xlsx workbook.")
            return redirect("admin_department")

        imported_faculty = 0
        imported_heads = 0

        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                code = str(sheet_name).strip().upper()

                # skip non-department sheets
                if code in ["HEAD", "OFFICES"]:
                    continue

                ws = wb[sheet_name]

                department_name = DEPARTMENT_MAP.get(code, str(sheet_name).strip())

                department, _ = Department.objects.get_or_create(
                    code=code,
                    defaults={"name": department_name},
                )

                department.name = department_name
                department.save()

                imported_faculty += _replace_faculty_from_department_sheet(
                    ws, department, selected_schedule
                )

            # import OFFICES sheet: OCD / ADAA
            if "OFFICES" in wb.sheetnames:
                ws = wb["OFFICES"]

                EvaluationOfficer.objects.filter(schedule=selected_schedule).delete()

                headers = [cell.value for cell in ws[1]]

                role_index = _get_header_index(headers, "ROLE")
                name_index = _get_header_index(headers, "NAME", "FULL NAME")
                email_index = _get_header_index(headers, "GSFE EMAIL", "EMAIL", "EMAIL ADDRESS")
                rank_index = _get_header_index(headers, "RANK", "ACADEMIC RANK", "POSITION")
                course_index = _get_header_index(headers, "COURSE", "COURSE CODE", "COURSE CODE/TITLE")
                program_year_index = _get_header_index(headers, "PROGRAM_YEAR", "PROGRAM YEAR", "YEAR LEVEL")

                for row in ws.iter_rows(min_row=2, values_only=True):
                    role = _get_row_value(row, role_index).upper()
                    officer_name = _get_row_value(row, name_index)
                    officer_email = _get_row_value(row, email_index)

                    if not role or not officer_name or not officer_email:
                        continue

                    if role not in ["OCD", "ADAA"]:
                        continue

                    EvaluationOfficer.objects.update_or_create(
                        schedule=selected_schedule,
                        role=role,
                        defaults={
                            "name": officer_name,
                            "email": officer_email,
                            "academic_rank": _get_row_value(row, rank_index),
                            "course": _get_row_value(row, course_index),
                            "program_year": _get_row_value(row, program_year_index),
                        },
                    )

            # import HEAD sheet
            if "HEAD" in wb.sheetnames:
                ws = wb["HEAD"]

                headers = [cell.value for cell in ws[1]]

                name_index = _get_header_index(headers, "NAME", "FULL NAME", "HEAD NAME")
                email_index = _get_header_index(headers, "GSFE EMAIL", "EMAIL", "EMAIL ADDRESS")
                department_index = _get_header_index(headers, "DEPARTMENT", "DEPT")
                rank_index = _get_header_index(headers, "RANK", "ACADEMIC RANK", "POSITION")
                course_index = _get_header_index(headers, "COURSE", "COURSE CODE", "COURSE CODE/TITLE")
                program_year_index = _get_header_index(headers, "PROGRAM_YEAR", "PROGRAM YEAR", "YEAR LEVEL")

                for row in ws.iter_rows(min_row=2, values_only=True):
                    head_name = _get_row_value(row, name_index)
                    head_email = _get_row_value(row, email_index)
                    dept_value = _get_row_value(row, department_index)

                    if not head_name or not dept_value:
                        continue

                    dept_key = dept_value.upper()

                    if dept_key in DEPARTMENT_MAP:
                        dept_code = dept_key
                        dept_name = DEPARTMENT_MAP[dept_key]
                    else:
                        matched_code = None

                        for code, full_name in DEPARTMENT_MAP.items():
                            if dept_value.lower() == full_name.lower():
                                matched_code = code
                                break

                        if matched_code:
                            dept_code = matched_code
                            dept_name = DEPARTMENT_MAP[matched_code]
                        else:
                            dept_code = dept_value.upper().replace(" ", "_")
                            dept_name = dept_value

                    department, _ = Department.objects.get_or_create(
                        code=dept_code,
                        defaults={"name": dept_name},
                    )

                    department.name = dept_name
                    department.save()

                    DepartmentHead.objects.update_or_create(
                        schedule=selected_schedule,
                        department=department,
                        defaults={
                            "name": head_name,
                            "email": head_email,
                            "academic_rank": _get_row_value(row, rank_index),
                            "course": _get_row_value(row, course_index),
                            "program_year": _get_row_value(row, program_year_index),
                        },
                    )

                    imported_heads += 1

        messages.success(
            request,
            f"Import complete: {imported_faculty} faculty and {imported_heads} department heads processed.",
        )
        return redirect("admin_department")

    schedules = EvaluationSchedule.objects.all().order_by("-start_datetime", "-created_at")
    selected_schedule_id = request.GET.get("schedule")

    display_schedule = None
    if selected_schedule_id:
        display_schedule = schedules.filter(id=selected_schedule_id).first()

    if not display_schedule:
        display_schedule = _get_open_schedule()

    if not display_schedule:
        display_schedule = _get_latest_schedule_with_uploaded_data()

    if not display_schedule:
        display_schedule = schedules.first()

    if display_schedule:
        departments = Department.objects.prefetch_related(
            Prefetch(
                "faculty_members",
                queryset=FacultyMember.objects.filter(schedule=display_schedule).order_by("name"),
            ),
            Prefetch(
                "heads",
                queryset=DepartmentHead.objects.filter(schedule=display_schedule).order_by("name"),
            ),
        ).order_by("name")

        total_faculty = FacultyMember.objects.filter(schedule=display_schedule).count()
    else:
        departments = Department.objects.prefetch_related(
            Prefetch("faculty_members", queryset=FacultyMember.objects.none()),
            Prefetch("heads", queryset=DepartmentHead.objects.none()),
        ).order_by("name")
        total_faculty = 0

    context = _admin_context(
        "department",
        {
            "departments": departments,
            "schedules": schedules,
            "total_departments": departments.count(),
            "total_faculty": total_faculty,
            "latest_department": departments.last(),
            "current_schedule": display_schedule,
        },
    )
    return render(request, "admin/admin_department.html", context)


@role_required("UITC", "ADAA")
def admin_manage(request):
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "save_schedule":
            schedule_id = request.POST.get("schedule_id")
            title = (request.POST.get("title") or "").strip()
            academic_year = (request.POST.get("academic_year") or "").strip()
            semester = (request.POST.get("semester") or "").strip()
            notes = (request.POST.get("notes") or "").strip()
            start_raw = request.POST.get("start_datetime")
            end_raw = request.POST.get("end_datetime")

            start_datetime = _parse_datetime_local(start_raw)
            end_datetime = _parse_datetime_local(end_raw)

            if not all([title, academic_year, semester, start_datetime, end_datetime]):
                messages.error(request, "Please complete all required schedule fields.")
                return redirect("admin_manage")

            if end_datetime <= start_datetime:
                messages.error(request, "Closing date and time must be later than the opening date and time.")
                return redirect("admin_manage")

            # exact duplicate check
            existing = EvaluationSchedule.objects.filter(
                title=title,
                academic_year=academic_year,
                semester=semester,
                start_datetime=start_datetime,
                end_datetime=end_datetime,
            )

            if schedule_id:
                existing = existing.exclude(id=schedule_id)

            if existing.exists():
                messages.error(
                    request,
                    "This schedule already exists. Please create another schedule with different details."
                )
                return redirect("admin_manage")

            # ONLY ONE OPEN SCHEDULE AT A TIME
            # This only blocks if the schedule being saved would become Open right now.
            now = timezone.localtime(timezone.now())

            submitted_schedule_will_be_open_now = (
                start_datetime <= now <= end_datetime
            )

            if submitted_schedule_will_be_open_now:
                existing_schedule = EvaluationSchedule.objects.filter(
                    start_datetime__lte=now,
                    end_datetime__gte=now
                )

                if schedule_id:
                    existing_schedule = existing_schedule.exclude(id=schedule_id)

                if existing_schedule.exists():
                    messages.error(
                        request,
                        "Only one evaluation schedule can be open at a time."
                    )
                    return redirect("admin_manage")

            if schedule_id:
                schedule = get_object_or_404(EvaluationSchedule, id=schedule_id)
                schedule.title = title
                schedule.academic_year = academic_year
                schedule.semester = semester
                schedule.start_datetime = start_datetime
                schedule.end_datetime = end_datetime
                schedule.notes = notes
                schedule.save()
                messages.success(request, "Evaluation schedule updated successfully.")
                return redirect("admin_manage")
            else:
                schedule = EvaluationSchedule.objects.create(
                    title=title,
                    academic_year=academic_year,
                    semester=semester,
                    start_datetime=start_datetime,
                    end_datetime=end_datetime,
                    notes=notes,
                )
                messages.success(request, "Evaluation schedule created successfully.")
                return redirect(f"{reverse('admin_department')}?schedule={schedule.id}")

        elif action == "delete_schedule":
            schedule_id = request.POST.get("schedule_id")
            schedule = get_object_or_404(EvaluationSchedule, id=schedule_id)
            schedule.delete()
            messages.success(request, "Evaluation schedule deleted successfully.")
            return redirect("admin_manage")

    schedules = EvaluationSchedule.objects.all().order_by("-start_datetime", "-created_at")
    now = timezone.localtime(timezone.now())

    context = _admin_context(
        "manage",
        {
            "schedules": schedules,
            "total_periods": schedules.count(),
            "open_periods": sum(1 for s in schedules if s.computed_status == "Open"),
            "closed_periods": sum(1 for s in schedules if s.computed_status == "Closed"),
            "current_time": now,
        },
    )
    return render(request, "admin/admin_manage.html", context)

@role_required("UITC", "ADAA")
def add_department(request):
    if request.method != "POST":
        return redirect("admin_department")

    code = (request.POST.get("code") or "").strip().upper()
    name = (request.POST.get("name") or "").strip()
    head_name = (request.POST.get("head_name") or "").strip()
    head_email = (request.POST.get("head_email") or "").strip()
    faculty_file = request.FILES.get("faculty_file")
    schedule_id = request.POST.get("schedule_id")

    if not schedule_id:
        messages.error(request, "Please select an evaluation schedule.")
        return redirect("admin_department")

    selected_schedule = get_object_or_404(EvaluationSchedule, id=schedule_id)

    if not code or not name:
        messages.error(request, "Department code and name are required.")
        return redirect(f"{reverse('admin_department')}?schedule={selected_schedule.id}")

    department, created = Department.objects.get_or_create(
        code=code,
        defaults={"name": name},
    )

    # If the department already exists, update its name instead of blocking it.
    department.name = name
    department.save()

    if head_name:
        DepartmentHead.objects.update_or_create(
            schedule=selected_schedule,
            department=department,
            defaults={
                "name": head_name,
                "email": head_email or "",
            },
        )

    if faculty_file:
        try:
            count = _replace_faculty_from_uploaded_file(
                faculty_file,
                department,
                selected_schedule
            )

            if created:
                messages.success(request, f"Department added successfully with {count} faculty members.")
            else:
                messages.success(request, f"Department already existed. Faculty list for this schedule was updated with {count} records.")

        except Exception as e:
            messages.warning(request, f"Department saved, but faculty file could not be processed: {str(e)}")
            return redirect(f"{reverse('admin_department')}?schedule={selected_schedule.id}")
    else:
        if created:
            messages.success(request, "Department added successfully.")
        else:
            messages.success(request, "Department already existed. Department details were updated for the selected schedule.")

    return redirect(f"{reverse('admin_department')}?schedule={selected_schedule.id}")


@role_required("UITC", "ADAA")
def update_department(request, dept_id):
    if request.method != "POST":
        return redirect("admin_department")

    department = get_object_or_404(Department, id=dept_id)

    code = (request.POST.get("code") or "").strip().upper()
    name = (request.POST.get("name") or "").strip()
    head_name = (request.POST.get("head_name") or "").strip()
    head_email = (request.POST.get("head_email") or "").strip()
    faculty_file = request.FILES.get("faculty_file")
    schedule_id = request.POST.get("schedule_id")

    if not schedule_id:
        messages.error(request, "Please select an evaluation schedule.")
        return redirect("admin_department")

    selected_schedule = get_object_or_404(EvaluationSchedule, id=schedule_id)
    redirect_url = f"{reverse('admin_department')}?schedule={selected_schedule.id}"

    if not code or not name:
        messages.error(request, "Department code and name are required.")
        return redirect(redirect_url)

    existing_department = Department.objects.filter(code=code).exclude(id=department.id).first()
    if existing_department:
        messages.error(request, f"Department code '{code}' is already used by another department.")
        return redirect(redirect_url)

    department.code = code
    department.name = name
    department.save()

    if head_name:
        DepartmentHead.objects.update_or_create(
            schedule=selected_schedule,
            department=department,
            defaults={
                "name": head_name,
                "email": head_email or "",
            },
        )
    else:
        DepartmentHead.objects.filter(
            schedule=selected_schedule,
            department=department
        ).delete()

    if faculty_file:
        try:
            count = _replace_faculty_from_uploaded_file(
                faculty_file,
                department,
                selected_schedule
            )
            messages.success(
                request,
                f"Department updated successfully. Faculty list replaced with {count} records."
            )
        except Exception as e:
            messages.warning(
                request,
                f"Department updated, but faculty file could not be processed: {str(e)}"
            )
            return redirect(redirect_url)
    else:
        messages.success(request, "Department updated successfully.")

    return redirect(redirect_url)


@role_required("UITC", "ADAA")
def delete_department(request, dept_id):
    if request.method != "POST":
        return redirect("admin_department")

    department = get_object_or_404(Department, id=dept_id)
    department_name = department.name
    department.delete()

    messages.success(request, f"Department '{department_name}' was deleted successfully.")
    return redirect("admin_department")

@role_required("UITC", "ADAA", "STAFF")
def admin_results_summary(request):
    schedules = EvaluationSchedule.objects.all().order_by("-start_datetime", "-created_at")
    selected_schedule_id = request.GET.get("schedule")

    selected_schedule = None
    if selected_schedule_id:
        selected_schedule = schedules.filter(id=selected_schedule_id).first()

    if not selected_schedule:
        selected_schedule = _get_latest_schedule_with_submitted_evaluations()

    if not selected_schedule:
        selected_schedule = _get_latest_schedule_with_uploaded_data()

    results = []

    if not selected_schedule:
        context = _admin_context(
            "results_summary",
            {
                "faculty_results": [],
                "departments": [],
                "total_faculty_count": 0,
                "highest_average_grade": 0,
                "lowest_average_grade": 0,
                "overall_faculty_average": 0,
                "selected_schedule": None,
                "academic_years": [],
                "semesters": [],
                "selected_academic_year": "",
                "selected_semester": "",
                "schedules": schedules,
            },
        )
        return render(request, "admin/admin_overall.html", context)

    grouped_results = {}

    def add_evaluation_to_group(
        grouped,
        result_type,
        schedule_obj,
        target_id,
        target_name,
        target_department,
        evaluator_name,
        evaluator_department,
        average_score,
        total_score,
        comments,
        submitted_at,
        responses,
    ):
        schedule_label = ""
        schedule_key = "no-schedule"

        if schedule_obj:
            schedule_key = str(schedule_obj.id)
            schedule_label = f"{schedule_obj.academic_year} | {schedule_obj.semester} | {schedule_obj.title}"

        group_key = f"{result_type}-{schedule_key}-{target_id}"

        if group_key not in grouped:
            grouped[group_key] = {
                "id": target_id,
                "result_type": result_type,
                "name": target_name,
                "department": target_department,
                "schedule_label": schedule_label,
                "academic_year": schedule_obj.academic_year if schedule_obj else "",
                "semester": schedule_obj.semester if schedule_obj else "",
                "title": schedule_obj.title if schedule_obj else "",
                "evaluators": [],
                "section_values": defaultdict(list),
                "overall_values": [],
                "total_scores": [],
                "computed_ratings": [],
            }

        section_groups = defaultdict(list)
        detailed_answers = defaultdict(list)

        for response in responses:
            section_key = (response.section_code or "").strip()
            section_name = (response.section_name or "").strip() or "Unnamed Section"
            rating_value = float(response.rating or 0)

            if section_key:
                section_groups[section_key].append(rating_value)

            detailed_answers[section_name].append({
                "question_number": response.question_number,
                "question_text": response.question_text or f"Question {response.question_number}",
                "rating": rating_value,
            })

        evaluator_sections = {
            "management_teaching_learning": 0,
            "content_knowledge_pedagogy_technology": 0,
            "commitment_transparency": 0,
        }

        for section_key, values in section_groups.items():
            evaluator_sections[section_key] = (sum(values) / len(values)) if values else 0
            grouped[group_key]["section_values"][section_key].append(evaluator_sections[section_key])

        evaluator_average_score = float(average_score or 0)
        evaluator_total_score = float(total_score or 0)
        evaluator_computed_rating = ((evaluator_total_score / 75) * 100) if evaluator_total_score else 0

        grouped[group_key]["evaluators"].append({
            "evaluator_name": evaluator_name or "Unknown Evaluator",
            "evaluator_department": evaluator_department or "",
            "average_score": evaluator_average_score,
            "total_score": evaluator_total_score,
            "computed_rating": evaluator_computed_rating,
            "submitted_at": submitted_at,
            "comments": comments or "",
            "sections": evaluator_sections,
            "detailed_answers": dict(detailed_answers),
        })

        grouped[group_key]["overall_values"].append(evaluator_average_score)
        grouped[group_key]["total_scores"].append(evaluator_total_score)
        grouped[group_key]["computed_ratings"].append(evaluator_computed_rating)

    # =========================
    # FACULTY RESULTS
    # =========================
    faculty_evaluations = (
        FacultyEvaluation.objects
        .filter(status="submitted", schedule=selected_schedule)
        .select_related(
            "evaluatee_faculty__department",
            "evaluator_head__department",
            "schedule",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(FacultyEvaluationResponse),
            )
        )
        .order_by("evaluatee_name", "evaluator_name", "submitted_at")
    )

    for evaluation in faculty_evaluations:
        add_evaluation_to_group(
            grouped=grouped_results,
            result_type="faculty",
            schedule_obj=evaluation.schedule,
            target_id=evaluation.evaluatee_faculty_id,
            target_name=evaluation.evaluatee_name,
            target_department=evaluation.evaluatee_department,
            evaluator_name=evaluation.evaluator_name,
            evaluator_department=evaluation.evaluator_department,
            average_score=evaluation.average_score,
            total_score=evaluation.total_score,
            comments=evaluation.comments,
            submitted_at=evaluation.submitted_at,
            responses=evaluation.responses.all(),
        )

# =========================
# HEAD RESULTS (ADAA -> Heads)
# =========================
    head_evaluations = (
        HeadEvaluation.objects
        .filter(status="submitted", schedule=selected_schedule)
        .select_related(
            "evaluatee_head__department",
            "evaluator_officer",
            "schedule",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(HeadEvaluationResponse),
            )
        )
        .order_by("evaluatee_name", "evaluator_name", "submitted_at")
    )

    for evaluation in head_evaluations:
        add_evaluation_to_group(
            grouped=grouped_results,
            result_type="head",
            schedule_obj=evaluation.schedule,
            target_id=evaluation.evaluatee_head_id,
            target_name=evaluation.evaluatee_name,
            target_department=evaluation.evaluatee_department or (
                evaluation.evaluatee_head.department.name if evaluation.evaluatee_head and evaluation.evaluatee_head.department else ""
            ),
            evaluator_name=evaluation.evaluator_name,
            evaluator_department="Office of the ADAA",
            average_score=evaluation.average_score,
            total_score=evaluation.total_score,
            comments=evaluation.comments,
            submitted_at=evaluation.submitted_at,
            responses=evaluation.responses.all(),
        )

    # =========================
    # OFFICE RESULTS (OCD -> ADAA)
    # =========================
    office_evaluations = (
        OfficeEvaluation.objects
        .filter(status="submitted", schedule=selected_schedule)
        .select_related(
            "evaluatee_officer",
            "evaluator_officer",
            "schedule",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(OfficeEvaluationResponse),
            )
        )
        .order_by("evaluatee_name", "evaluator_name", "submitted_at")
    )

    for evaluation in office_evaluations:
        target_department = (
            "Office of the ADAA"
            if (evaluation.evaluatee_role or "").upper() == "ADAA"
            else "Office of the Campus Director"
        )

        evaluator_department = (
            "Office of the Campus Director"
            if (evaluation.evaluator_role or "").upper() == "OCD"
            else "Office of the ADAA"
        )

        add_evaluation_to_group(
            grouped=grouped_results,
            result_type="office",
            schedule_obj=evaluation.schedule,
            target_id=evaluation.evaluatee_officer_id,
            target_name=evaluation.evaluatee_name,
            target_department=target_department,
            evaluator_name=evaluation.evaluator_name,
            evaluator_department=evaluator_department,
            average_score=evaluation.average_score,
            total_score=evaluation.total_score,
            comments=evaluation.comments,
            submitted_at=evaluation.submitted_at,
            responses=evaluation.responses.all(),
        )

    overall_list = []
    results = []

    for index, group in enumerate(grouped_results.values(), start=1):
        evaluator_count = len(group["evaluators"])

        average_score = (sum(group["overall_values"]) / evaluator_count) if evaluator_count else 0
        average_total_score = (sum(group["total_scores"]) / evaluator_count) if evaluator_count else 0
        computed_rating = (sum(group["computed_ratings"]) / evaluator_count) if evaluator_count else 0

        section_averages = {
            "management_teaching_learning": 0,
            "content_knowledge_pedagogy_technology": 0,
            "commitment_transparency": 0,
        }

        for section_key, values in group["section_values"].items():
            section_averages[section_key] = (sum(values) / len(values)) if values else 0

        overall_list.append(computed_rating)

        results.append({
            "num": index,
            "id": group["id"],
            "result_type": group["result_type"],
            "name": group["name"],
            "department": group["department"],
            "position": (
                "Faculty Member" if group["result_type"] == "faculty"
                else "Department Head" if group["result_type"] == "head"
                else "Evaluation Officer"
            ),
            "evaluator_count": evaluator_count,
            "average_score": average_score,
            "average_total_score": average_total_score,
            "computed_rating": computed_rating,
            "sections": section_averages,
            "evaluators": group["evaluators"],
            "academic_year": group["academic_year"],
            "semester": group["semester"],
            "title": group["title"],
            "schedule_label": group["schedule_label"],
        })

    type_order = {"office": 0, "head": 1, "faculty": 2}
    results.sort(key=lambda item: (type_order.get(item["result_type"], 99), item["name"].lower()))

    departments = sorted({item["department"] for item in results if item["department"]})

    context = _admin_context(
        "results_summary",
        {
            "faculty_results": results,
            "departments": departments,
            "total_faculty_count": len(results),
            "highest_average_grade": max(overall_list) if overall_list else 0,
            "lowest_average_grade": min(overall_list) if overall_list else 0,
            "overall_faculty_average": (sum(overall_list) / len(overall_list)) if overall_list else 0,
            "selected_schedule": selected_schedule,
            "academic_years": [],
            "semesters": [],
            "selected_academic_year": "",
            "selected_semester": "",
            "schedules": schedules,
        },
    )

    return render(request, "admin/admin_overall.html", context)


# ==========================================================
# EXPORT FULL EVALUATION RESULTS TO EXCEL
# ==========================================================

SECTION_CODE_MAP = {
    "management_teaching_learning": "A. Management of Teaching and Learning",
    "content_knowledge_pedagogy_technology": "B. Content Knowledge, Pedagogy, and Technology",
    "commitment_transparency": "C. Commitment and Transparency",
}

QUESTION_BANK = {
    1: "Comes to class on time.",
    2: "Submits updated syllabus, grade sheets, and other required reports on time.",
    3: "Maximizes the allocated time or learning hours effectively.",
    4: "Provides appropriate learning activities that facilitate students' critical thinking and creativity.",
    5: "Guides students to learn on their own, reflect on new ideas and experiences, and make decisions in accomplishing given tasks.",
    6: "Communicates constructive feedback to students for their academic growth.",
    7: "Demonstrates extensive and broad knowledge of the subject or course.",
    8: "Simplifies complex ideas in the lesson for ease of understanding.",
    9: "Integrates contemporary issues and developments in the discipline and/or daily life activities in the syllabus.",
    10: "Promotes active learning and student engagement by using appropriate teaching and learning resources including ICT tools and platforms.",
    11: "Uses appropriate assessments such as projects, exams, quizzes, and assignments aligned with the learning outcomes.",
    12: "Recognizes and values the unique diversity and individual differences among students.",
    13: "Assists students with their learning challenges during consultation hours.",
    14: "Provides immediate feedback on student outputs and performance.",
    15: "Provides transparent and clear criteria in rating student performance.",
}

SECTION_CONTINUOUS_NUMBER_MAP = {
    "management_teaching_learning": {
        1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6,
    },
    "content_knowledge_pedagogy_technology": {
        1: 7, 2: 8, 3: 9, 4: 10, 5: 11,
    },
    "commitment_transparency": {
        1: 12, 2: 13, 3: 14, 4: 15,
    },
}


def _safe_float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0


def _excel_datetime(value):
    if not value:
        return ""

    try:
        return timezone.localtime(value).strftime("%Y-%m-%d %I:%M %p")
    except Exception:
        return str(value)


def _normalize_eval_section_code(section_code="", section_name=""):
    section_code = (section_code or "").strip()
    section_name = (section_name or "").strip().lower()

    if section_code:
        return section_code

    if "management" in section_name and "teaching" in section_name:
        return "management_teaching_learning"

    if "content" in section_name or "pedagogy" in section_name or "technology" in section_name:
        return "content_knowledge_pedagogy_technology"

    if "commitment" in section_name or "transparency" in section_name:
        return "commitment_transparency"

    return ""


def _get_continuous_question_number(section_code, question_number):
    try:
        question_number = int(question_number)
    except (TypeError, ValueError):
        return None

    if section_code in SECTION_CONTINUOUS_NUMBER_MAP:
        return SECTION_CONTINUOUS_NUMBER_MAP[section_code].get(question_number, question_number)

    return question_number


def _get_response_data(responses):
    """
    Returns:
    - question_ratings: {1: 5, 2: 4, ...}
    - section_averages: {'management_teaching_learning': 4.50, ...}
    """
    question_ratings = {}
    section_values = defaultdict(list)

    for response in responses:
        section_code = _normalize_eval_section_code(
            response.section_code,
            response.section_name
        )

        rating = _safe_float(response.rating)
        continuous_question_number = _get_continuous_question_number(
            section_code,
            response.question_number
        )

        if continuous_question_number:
            question_ratings[continuous_question_number] = rating

        if section_code:
            section_values[section_code].append(rating)

    section_averages = {
        "management_teaching_learning": 0,
        "content_knowledge_pedagogy_technology": 0,
        "commitment_transparency": 0,
    }

    for section_code, values in section_values.items():
        section_averages[section_code] = round(sum(values) / len(values), 2) if values else 0

    return question_ratings, section_averages


def _style_excel_sheet(ws):
    header_fill = PatternFill("solid", fgColor="981B2E")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    ws.freeze_panes = "A2"

    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))

        ws.column_dimensions[col_letter].width = min(max_length + 3, 45)


def _get_selected_export_schedule(request):
    schedules = EvaluationSchedule.objects.all().order_by("-start_datetime", "-created_at")
    selected_schedule_id = request.GET.get("schedule")

    selected_schedule = None

    if selected_schedule_id:
        selected_schedule = schedules.filter(id=selected_schedule_id).first()

    if not selected_schedule:
        selected_schedule = _get_latest_schedule_with_submitted_evaluations()

    if not selected_schedule:
        selected_schedule = _get_latest_schedule_with_uploaded_data()

    return selected_schedule


def _get_department_head_name(schedule, department):
    if not schedule or not department:
        return ""

    head = DepartmentHead.objects.filter(
        schedule=schedule,
        department=department
    ).first()

    return head.name if head else ""


@role_required("UITC", "ADAA", "STAFF")
def export_results_excel(request):
    selected_schedule = _get_selected_export_schedule(request)

    if not selected_schedule:
        messages.error(request, "No evaluation schedule found to export.")
        return redirect("admin_results_summary")

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"

    detail_ws = wb.create_sheet("Detailed Ratings")
    guide_ws = wb.create_sheet("Question Guide")

    # ======================================================
    # QUESTION GUIDE SHEET
    # ======================================================
    guide_ws.append([
        "Question No.",
        "Benchmark Statement",
    ])

    for question_no, question_text in QUESTION_BANK.items():
        guide_ws.append([
            f"Q{question_no}",
            question_text,
        ])

    # ======================================================
    # SUMMARY SHEET
    # One row per evaluatee.
    # ======================================================
    summary_headers = [
        "No.",
        "Academic Year",
        "Semester",
        "Evaluation Title",
        "Evaluatee Type",
        "Evaluatee Name",
        "Department / Office",
        "Department Head / Supervisor",
        "Evaluator Count",
        "Average Total Score",
        "Average Score",
        "Computed Rating (%)",
        "A. Management of Teaching and Learning",
        "B. Content Knowledge, Pedagogy, and Technology",
        "C. Commitment and Transparency",
    ]

    summary_ws.append(summary_headers)

    # ======================================================
    # DETAILED RATINGS SHEET
    # One row per evaluator submission.
    # ======================================================
    detail_headers = [
        "No.",
        "Academic Year",
        "Semester",
        "Evaluation Title",
        "Evaluatee Type",
        "Evaluatee Name",
        "Department / Office",
        "Department Head / Supervisor",
        "Evaluator Type",
        "Evaluator Name",
        "Evaluator Department / Office",
        "Submitted At",
        "Total Score",
        "Average Score",
        "Computed Rating (%)",
        "A. Management of Teaching and Learning",
        "B. Content Knowledge, Pedagogy, and Technology",
        "C. Commitment and Transparency",
    ]

    for question_no in range(1, 16):
        detail_headers.append(f"Q{question_no} Rating")

    detail_headers.append("Comments")

    detail_ws.append(detail_headers)

    summary_groups = {}

    def add_detail_row(
        result_type,
        evaluatee_type,
        evaluatee_name,
        department_or_office,
        department_head_or_supervisor,
        evaluator_type,
        evaluator_name,
        evaluator_department,
        evaluation,
        responses,
    ):
        question_ratings, section_averages = _get_response_data(responses)

        total_score = _safe_float(evaluation.total_score)
        average_score = _safe_float(evaluation.average_score)
        computed_rating = round((total_score / 75) * 100, 2) if total_score else 0

        row_no = detail_ws.max_row

        detail_row = [
            row_no,
            selected_schedule.academic_year,
            selected_schedule.semester,
            selected_schedule.title,
            evaluatee_type,
            evaluatee_name,
            department_or_office,
            department_head_or_supervisor,
            evaluator_type,
            evaluator_name,
            evaluator_department,
            _excel_datetime(evaluation.submitted_at),
            round(total_score, 2),
            round(average_score, 2),
            computed_rating,
            section_averages.get("management_teaching_learning", 0),
            section_averages.get("content_knowledge_pedagogy_technology", 0),
            section_averages.get("commitment_transparency", 0),
        ]

        for question_no in range(1, 16):
            detail_row.append(question_ratings.get(question_no, ""))

        detail_row.append(evaluation.comments or "")

        detail_ws.append(detail_row)

        group_key = f"{result_type}-{evaluatee_name}-{department_or_office}"

        if group_key not in summary_groups:
            summary_groups[group_key] = {
                "evaluatee_type": evaluatee_type,
                "evaluatee_name": evaluatee_name,
                "department_or_office": department_or_office,
                "department_head_or_supervisor": department_head_or_supervisor,
                "total_scores": [],
                "average_scores": [],
                "computed_ratings": [],
                "section_values": {
                    "management_teaching_learning": [],
                    "content_knowledge_pedagogy_technology": [],
                    "commitment_transparency": [],
                },
            }

        summary_groups[group_key]["total_scores"].append(total_score)
        summary_groups[group_key]["average_scores"].append(average_score)
        summary_groups[group_key]["computed_ratings"].append(computed_rating)

        for section_code in summary_groups[group_key]["section_values"]:
            summary_groups[group_key]["section_values"][section_code].append(
                section_averages.get(section_code, 0)
            )

    # ======================================================
    # FACULTY RESULTS
    # Faculty evaluated by Department Heads.
    # ======================================================
    faculty_evaluations = (
        FacultyEvaluation.objects
        .filter(schedule=selected_schedule, status="submitted")
        .select_related(
            "schedule",
            "evaluatee_faculty__department",
            "evaluator_head__department",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(FacultyEvaluationResponse),
            )
        )
        .order_by("evaluatee_name", "evaluator_name", "submitted_at")
    )

    for evaluation in faculty_evaluations:
        evaluatee_department = (
            evaluation.evaluatee_department
            or (
                evaluation.evaluatee_faculty.department.name
                if evaluation.evaluatee_faculty and evaluation.evaluatee_faculty.department
                else ""
            )
        )

        department_head_name = (
            evaluation.evaluator_head.name
            if evaluation.evaluator_head
            else evaluation.evaluator_name
        )

        add_detail_row(
            result_type="faculty",
            evaluatee_type="Faculty Member",
            evaluatee_name=evaluation.evaluatee_name or (
                evaluation.evaluatee_faculty.name if evaluation.evaluatee_faculty else ""
            ),
            department_or_office=evaluatee_department,
            department_head_or_supervisor=department_head_name,
            evaluator_type="Department Head",
            evaluator_name=evaluation.evaluator_name,
            evaluator_department=evaluation.evaluator_department,
            evaluation=evaluation,
            responses=evaluation.responses.all(),
        )

    # ======================================================
    # HEAD RESULTS
    # Department Heads evaluated by ADAA.
    # ======================================================
    head_evaluations = (
        HeadEvaluation.objects
        .filter(schedule=selected_schedule, status="submitted")
        .select_related(
            "schedule",
            "evaluatee_head__department",
            "evaluator_officer",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(HeadEvaluationResponse),
            )
        )
        .order_by("evaluatee_name", "evaluator_name", "submitted_at")
    )

    for evaluation in head_evaluations:
        evaluatee_department = (
            evaluation.evaluatee_department
            or (
                evaluation.evaluatee_head.department.name
                if evaluation.evaluatee_head and evaluation.evaluatee_head.department
                else ""
            )
        )

        add_detail_row(
            result_type="head",
            evaluatee_type="Department Head",
            evaluatee_name=evaluation.evaluatee_name or (
                evaluation.evaluatee_head.name if evaluation.evaluatee_head else ""
            ),
            department_or_office=evaluatee_department,
            department_head_or_supervisor=evaluation.evaluatee_name,
            evaluator_type=evaluation.evaluator_role or "ADAA",
            evaluator_name=evaluation.evaluator_name,
            evaluator_department="Office of the ADAA",
            evaluation=evaluation,
            responses=evaluation.responses.all(),
        )

    # ======================================================
    # OFFICE RESULTS
    # ADAA evaluated by OCD.
    # ======================================================
    office_evaluations = (
        OfficeEvaluation.objects
        .filter(schedule=selected_schedule, status="submitted")
        .select_related(
            "schedule",
            "evaluatee_officer",
            "evaluator_officer",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(OfficeEvaluationResponse),
            )
        )
        .order_by("evaluatee_name", "evaluator_name", "submitted_at")
    )

    for evaluation in office_evaluations:
        evaluatee_office = (
            "Office of the ADAA"
            if (evaluation.evaluatee_role or "").upper() == "ADAA"
            else "Office of the Campus Director"
        )

        evaluator_office = (
            "Office of the Campus Director"
            if (evaluation.evaluator_role or "").upper() == "OCD"
            else "Office of the ADAA"
        )

        add_detail_row(
            result_type="office",
            evaluatee_type="Evaluation Officer",
            evaluatee_name=evaluation.evaluatee_name,
            department_or_office=evaluatee_office,
            department_head_or_supervisor="",
            evaluator_type=evaluation.evaluator_role or "OCD",
            evaluator_name=evaluation.evaluator_name,
            evaluator_department=evaluator_office,
            evaluation=evaluation,
            responses=evaluation.responses.all(),
        )

    # ======================================================
    # BUILD SUMMARY ROWS
    # ======================================================
    for index, group in enumerate(summary_groups.values(), start=1):
        evaluator_count = len(group["total_scores"])

        avg_total = (
            sum(group["total_scores"]) / evaluator_count
            if evaluator_count
            else 0
        )

        avg_score = (
            sum(group["average_scores"]) / evaluator_count
            if evaluator_count
            else 0
        )

        avg_computed = (
            sum(group["computed_ratings"]) / evaluator_count
            if evaluator_count
            else 0
        )

        section_a = group["section_values"]["management_teaching_learning"]
        section_b = group["section_values"]["content_knowledge_pedagogy_technology"]
        section_c = group["section_values"]["commitment_transparency"]

        summary_ws.append([
            index,
            selected_schedule.academic_year,
            selected_schedule.semester,
            selected_schedule.title,
            group["evaluatee_type"],
            group["evaluatee_name"],
            group["department_or_office"],
            group["department_head_or_supervisor"],
            evaluator_count,
            round(avg_total, 2),
            round(avg_score, 2),
            round(avg_computed, 2),
            round(sum(section_a) / len(section_a), 2) if section_a else 0,
            round(sum(section_b) / len(section_b), 2) if section_b else 0,
            round(sum(section_c) / len(section_c), 2) if section_c else 0,
        ])

    # ======================================================
    # STYLE SHEETS
    # ======================================================
    _style_excel_sheet(summary_ws)
    _style_excel_sheet(detail_ws)
    _style_excel_sheet(guide_ws)

    safe_filename = slugify(
        f"evaluation-results-{selected_schedule.academic_year}-{selected_schedule.semester}-{selected_schedule.title}"
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{safe_filename}.xlsx"'

    wb.save(response)
    return response


def admin_overall(request):
    return admin_results_summary(request)


def admin_login(request):
    open_schedule = _get_open_schedule()
    portal_closed = open_schedule is None

    if request.method == "POST":
        login_type = (request.POST.get("login_type") or "").strip()
        
        if login_type == "admin":
            username = (request.POST.get("username") or "").strip()
            password = (request.POST.get("password") or "").strip()

            if not username or not password:
                request.session["login_modal"] = {
                    "type": "danger",
                    "message": "Please enter both username and password for admin login."
                }
                return redirect("admin_login")

            user = authenticate(request, username=username, password=password)

            if user is None:
                request.session["login_modal"] = {
                    "type": "danger",
                    "message": "Invalid username or password."
                }
                return redirect("admin_login")

            if not user.is_active:
                request.session["login_modal"] = {
                    "type": "danger",
                    "message": "This account is inactive."
                }
                return redirect("admin_login")

            if not user.is_staff:
                request.session["login_modal"] = {
                    "type": "danger",
                    "message": "You do not have admin access."
                }
                return redirect("admin_login")

            user_role = get_admin_role(user)

            if not user_role:
                request.session["login_modal"] = {
                    "type": "danger",
                    "message": "Your admin account has no assigned role. Please contact UITC."
                }
                return redirect("admin_login")

            login(request, user)
            return redirect(admin_landing_page(user))
        
        elif login_type == "head":
            email = (request.POST.get("email") or "").strip().lower()

            if not email:
                request.session["login_modal"] = {
                    "type": "danger",
                    "message": "Please enter your GSFE email address."
                }
                return redirect("admin_login")

            head = (
                DepartmentHead.objects
                .select_related("department", "schedule")
                .filter(email__iexact=email)
                .order_by("-schedule__start_datetime", "-id")
                .first()
            )

            faculty = (
                FacultyMember.objects
                .select_related("department", "schedule")
                .filter(email__iexact=email)
                .order_by("-schedule__start_datetime", "-id")
                .first()
            )

            if not head:
                if faculty:
                    request.session["login_modal"] = {
                        "type": "danger",
                        "message": "This account is registered as faculty only. Faculty members are not allowed to access the department head portal."
                    }
                else:
                    request.session["login_modal"] = {
                        "type": "danger",
                        "message": "This email is not registered as a department head in the evaluation system."
                    }
                return redirect("admin_login")

            signer = TimestampSigner(salt=LINK_SALT)
            token = signer.sign(str(head.id))

            verify_url = request.build_absolute_uri(
                reverse("verify_head_login_link", args=[token])
            )

            subject = "Department Head Portal Login Link"

            context = {
                "recipient_name": head.name,
                "head": head,
                "verify_url": verify_url,
                "expires_minutes": LOGIN_LINK_MAX_AGE // 60,
                "open_schedule": open_schedule,
            }

            text_body = (
                f"Hello {head.name},\n\n"
                f"Click the link below to access the Department Head Portal:\n\n"
                f"{verify_url}\n\n"
                f"This link will expire in {LOGIN_LINK_MAX_AGE // 60} minutes.\n"
                f"If you did not request this, please ignore this email."
            )

            html_body = render_to_string("head/email_head_portal_link.html", context)

            try:
                msg = EmailMultiAlternatives(
                    subject=subject,
                    body=text_body,
                    from_email=getattr(settings, "DEFAULT_FROM_EMAIL", None),
                    to=[head.email],
                )
                msg.attach_alternative(html_body, "text/html")
                msg.send()

                request.session["login_modal"] = {
                    "type": "success",
                    "message": f"A secure login link has been sent to {head.email}."
                }
            except Exception:
                request.session["login_modal"] = {
                    "type": "danger",
                    "message": "The login link could not be sent. Please check your email settings."
                }

            return redirect("admin_login")

        else:
            request.session["login_modal"] = {
                "type": "danger",
                "message": "Please choose a login type first."
            }
            return redirect("admin_login")

    login_modal = request.session.pop("login_modal", None)

    context = _admin_context(
        "login",
        {
            "login_modal": login_modal,
            "open_schedule": open_schedule,
            "portal_closed": portal_closed,
        },
    )
    return render(request, "admin/admin_login.html", context)

def admin_forgot_password(request):
    if request.method != "POST":
        return redirect("admin_login")

    email = (request.POST.get("forgot_email") or "").strip().lower()

    if not email:
        request.session["login_modal"] = {
            "type": "danger",
            "message": "Please enter your admin email address."
        }
        return redirect("admin_login")

    user = (
        User.objects
        .filter(email__iexact=email, is_staff=True, is_active=True)
        .order_by("id")
        .first()
    )

    if not user:
        request.session["login_modal"] = {
            "type": "danger",
            "message": "No active admin account is registered with that email address."
        }
        return redirect("admin_login")

    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)

    reset_url = request.build_absolute_uri(
        reverse("admin_reset_password", args=[uid, token])
    )

    subject = "Admin Password Reset Request"

    text_body = (
        f"Hello {user.username},\n\n"
        f"You requested to reset your admin password.\n\n"
        f"Click the link below to continue:\n\n"
        f"{reset_url}\n\n"
        f"If you did not request this, please ignore this email."
    )

    html_body = f"""
        <div style="font-family: Arial, sans-serif; line-height: 1.6; color: #1f2937;">
            <h2 style="color: #981b2e;">Admin Password Reset</h2>
            <p>Hello <strong>{user.username}</strong>,</p>
            <p>You requested to reset your admin password.</p>
            <p>
                <a href="{reset_url}"
                   style="display:inline-block;padding:12px 20px;background:#981b2e;color:#fff;text-decoration:none;border-radius:8px;font-weight:700;">
                   Reset Password
                </a>
            </p>
            <p>If the button does not work, copy and paste this link into your browser:</p>
            <p>{reset_url}</p>
            <p>If you did not request this, you may ignore this email.</p>
        </div>
    """

    try:
        msg = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=getattr(settings, "EMAIL_HOST_USER", None),
            to=[user.email],
        )
        msg.attach_alternative(html_body, "text/html")
        msg.send()

        request.session["login_modal"] = {
            "type": "success",
            "message": f"A password reset link has been sent to {user.email}."
        }
    except Exception:
        request.session["login_modal"] = {
            "type": "danger",
            "message": "Password reset email could not be sent. Please check your email settings."
        }

    return redirect("admin_login")


def admin_reset_password(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid, is_staff=True)
    except Exception:
        user = None

    if user is None or not default_token_generator.check_token(user, token):
        return render(request, "admin/admin_reset_password.html", {
            "validlink": False
        })

    if request.method == "POST":
        password1 = (request.POST.get("password1") or "").strip()
        password2 = (request.POST.get("password2") or "").strip()

        if not password1 or not password2:
            return render(request, "admin/admin_reset_password.html", {
                "validlink": True,
                "error_message": "Please fill in both password fields."
            })

        if password1 != password2:
            return render(request, "admin/admin_reset_password.html", {
                "validlink": True,
                "error_message": "Passwords do not match."
            })

        try:
            validate_password(password1, user=user)
        except ValidationError as e:
            return render(request, "admin/admin_reset_password.html", {
                "validlink": True,
                "error_message": " ".join(e.messages)
            })

        user.set_password(password1)
        user.save()

        request.session["login_modal"] = {
            "type": "success",
            "message": "Your password has been reset successfully. You may now log in."
        }
        return redirect("admin_login")

    return render(request, "admin/admin_reset_password.html", {
        "validlink": True
    })


@role_required("UITC", "ADAA", "STAFF")
def admin_past_evaluations(request):
    selected_schedule_id = request.GET.get("schedule")

    past_schedules = (
        EvaluationSchedule.objects
        .filter(end_datetime__lt=timezone.localtime(timezone.now()))
        .order_by("-start_datetime", "-created_at")
    )

    selected_schedule = None
    if selected_schedule_id:
        selected_schedule = past_schedules.filter(id=selected_schedule_id).first()

    if not selected_schedule:
        selected_schedule = past_schedules.first()

    faculty_history_results = []
    history_results = []

    if selected_schedule:
        grouped_results = {}

        def add_to_group(
            grouped,
            result_type,
            schedule_obj,
            group_key,
            target_id,
            target_name,
            target_department,
            evaluation,
            responses,
            evaluator_name,
            evaluator_department,
        ):
            schedule_label = ""
            schedule_id_value = None
            academic_year = ""
            semester = ""
            title = ""

            if schedule_obj:
                schedule_id_value = schedule_obj.id
                academic_year = schedule_obj.academic_year or ""
                semester = schedule_obj.semester or ""
                title = schedule_obj.title or ""
                schedule_label = f"{academic_year} | {semester} | {title}"

            if group_key not in grouped:
                grouped[group_key] = {
                    "id": target_id,
                    "result_type": result_type,
                    "name": target_name,
                    "department": target_department,
                    "schedule_id": schedule_id_value,
                    "schedule_label": schedule_label,
                    "academic_year": academic_year,
                    "semester": semester,
                    "title": title,
                    "evaluators": [],
                    "section_values": defaultdict(list),
                    "overall_values": [],
                    "total_scores": [],
                    "computed_ratings": [],
                }

            section_groups = defaultdict(list)
            detailed_answers = defaultdict(list)

            for response in responses:
                section_key = (response.section_code or "").strip()
                section_name = (response.section_name or "").strip() or "Unnamed Section"
                rating_value = float(response.rating or 0)

                if section_key:
                    section_groups[section_key].append(rating_value)

                detailed_answers[section_name].append({
                    "question_number": response.question_number,
                    "question_text": response.question_text or f"Question {response.question_number}",
                    "rating": rating_value,
                })

            evaluator_sections = {
                "management_teaching_learning": 0,
                "content_knowledge_pedagogy_technology": 0,
                "commitment_transparency": 0,
            }

            for section_key, ratings in section_groups.items():
                evaluator_sections[section_key] = round(sum(ratings) / len(ratings), 2) if ratings else 0
                grouped[group_key]["section_values"][section_key].append(evaluator_sections[section_key])

            evaluator_total_score = round(float(evaluation.total_score or 0), 2)
            evaluator_overall = float(evaluation.average_score or 0)
            evaluator_computed_rating = round((evaluator_total_score / 75) * 100, 2) if evaluator_total_score else 0

            grouped[group_key]["evaluators"].append({
                "evaluator_name": evaluator_name or "Unknown Evaluator",
                "evaluator_department": evaluator_department or "",
                "sections": evaluator_sections,
                "overall": evaluator_overall,
                "total_score": evaluator_total_score,
                "computed_rating": evaluator_computed_rating,
                "comments": evaluation.comments or "",
                "submitted_at": evaluation.submitted_at.strftime("%Y-%m-%d %H:%M") if evaluation.submitted_at else "",
                "detailed_answers": dict(detailed_answers),
            })

            grouped[group_key]["overall_values"].append(evaluator_overall)
            grouped[group_key]["total_scores"].append(evaluator_total_score)
            grouped[group_key]["computed_ratings"].append(evaluator_computed_rating)

        # =========================
        # FACULTY RESULTS
        # =========================
        faculty_evaluations = (
            FacultyEvaluation.objects
            .filter(schedule=selected_schedule, status="submitted")
            .select_related(
                "evaluatee_faculty__department",
                "evaluator_head__department",
                "schedule",
            )
            .prefetch_related(
                Prefetch(
                    "responses",
                    queryset=_ordered_response_queryset(FacultyEvaluationResponse),
                )
            )
            .order_by("evaluatee_name", "evaluator_name", "submitted_at")
        )

        for evaluation in faculty_evaluations:
            target_id = evaluation.evaluatee_faculty.id if evaluation.evaluatee_faculty else f"faculty-eval-{evaluation.id}"
            target_name = (
                evaluation.evaluatee_name
                or (evaluation.evaluatee_faculty.name if evaluation.evaluatee_faculty else "Unknown Faculty")
            )
            target_department = (
                evaluation.evaluatee_department
                or (
                    evaluation.evaluatee_faculty.department.name
                    if evaluation.evaluatee_faculty and evaluation.evaluatee_faculty.department
                    else ""
                )
            )

            add_to_group(
                grouped=grouped_results,
                result_type="faculty",
                schedule_obj=evaluation.schedule,
                group_key=f"faculty-{selected_schedule.id}-{target_id}",
                target_id=target_id,
                target_name=target_name,
                target_department=target_department,
                evaluation=evaluation,
                responses=evaluation.responses.all(),
                evaluator_name=evaluation.evaluator_name,
                evaluator_department=evaluation.evaluator_department,
            )

        # =========================
        # HEAD RESULTS
        # =========================
        head_evaluations = (
            HeadEvaluation.objects
            .filter(schedule=selected_schedule, status="submitted")
            .select_related(
                "evaluatee_head__department",
                "evaluator_officer",
                "schedule",
            )
            .prefetch_related(
                Prefetch(
                    "responses",
                    queryset=_ordered_response_queryset(HeadEvaluationResponse),
                )
            )
            .order_by("evaluatee_name", "evaluator_name", "submitted_at")
        )

        for evaluation in head_evaluations:
            target_id = evaluation.evaluatee_head.id if evaluation.evaluatee_head else f"head-eval-{evaluation.id}"
            target_name = (
                evaluation.evaluatee_name
                or (evaluation.evaluatee_head.name if evaluation.evaluatee_head else "Unknown Department Head")
            )
            target_department = (
                evaluation.evaluatee_department
                or (
                    evaluation.evaluatee_head.department.name
                    if evaluation.evaluatee_head and evaluation.evaluatee_head.department
                    else ""
                )
            )

            add_to_group(
                grouped=grouped_results,
                result_type="head",
                schedule_obj=evaluation.schedule,
                group_key=f"head-{selected_schedule.id}-{target_id}",
                target_id=target_id,
                target_name=target_name,
                target_department=target_department,
                evaluation=evaluation,
                responses=evaluation.responses.all(),
                evaluator_name=evaluation.evaluator_name,
                evaluator_department="Office of the ADAA",
            )

        # =========================
        # OFFICE RESULTS
        # =========================
        office_evaluations = (
            OfficeEvaluation.objects
            .filter(schedule=selected_schedule, status="submitted")
            .select_related(
                "evaluatee_officer",
                "evaluator_officer",
                "schedule",
            )
            .prefetch_related(
                Prefetch(
                    "responses",
                    queryset=_ordered_response_queryset(OfficeEvaluationResponse),
                )
            )
            .order_by("evaluatee_name", "evaluator_name", "submitted_at")
        )

        for evaluation in office_evaluations:
            target_id = evaluation.evaluatee_officer.id if evaluation.evaluatee_officer else f"office-eval-{evaluation.id}"
            target_name = evaluation.evaluatee_name or (
                evaluation.evaluatee_officer.name if evaluation.evaluatee_officer else "Unknown Officer"
            )

            target_department = (
                "Office of the ADAA"
                if (evaluation.evaluatee_role or "").upper() == "ADAA"
                else "Office of the Campus Director"
            )

            evaluator_department = (
                "Office of the Campus Director"
                if (evaluation.evaluator_role or "").upper() == "OCD"
                else "Office of the ADAA"
            )

            add_to_group(
                grouped=grouped_results,
                result_type="office",
                schedule_obj=evaluation.schedule,
                group_key=f"office-{selected_schedule.id}-{target_id}",
                target_id=target_id,
                target_name=target_name,
                target_department=target_department,
                evaluation=evaluation,
                responses=evaluation.responses.all(),
                evaluator_name=evaluation.evaluator_name,
                evaluator_department=evaluator_department,
            )

        all_history_results = []

        for _, item in grouped_results.items():
            section_averages = {
                "management_teaching_learning": 0,
                "content_knowledge_pedagogy_technology": 0,
                "commitment_transparency": 0,
            }

            for section_key, values in item["section_values"].items():
                section_averages[section_key] = round(sum(values) / len(values), 2) if values else 0

            overall_average = round(sum(item["overall_values"]) / len(item["overall_values"]), 3) if item["overall_values"] else 0
            average_total_score = round(sum(item["total_scores"]) / len(item["total_scores"]), 2) if item["total_scores"] else 0
            computed_rating = round(sum(item["computed_ratings"]) / len(item["computed_ratings"]), 2) if item["computed_ratings"] else 0

            all_history_results.append({
                "id": item["id"],
                "result_type": item["result_type"],
                "name": item["name"],
                "department": item["department"],
                "schedule_id": item["schedule_id"],
                "schedule_label": item["schedule_label"],
                "academic_year": item["academic_year"],
                "semester": item["semester"],
                "title": item["title"],
                "sections": section_averages,
                "overall": overall_average,
                "average_total_score": average_total_score,
                "computed_rating": computed_rating,
                "evaluator_count": len(item["evaluators"]),
                "evaluators": item["evaluators"],
            })

        type_order = {"office": 0, "head": 1, "faculty": 2}
        all_history_results.sort(key=lambda x: (type_order.get(x["result_type"], 99), str(x["name"]).lower()))

        faculty_history_results = [item for item in all_history_results if item["result_type"] == "faculty"]
        history_results = all_history_results

    context = _admin_context("past_evaluations", {
        "past_schedules": past_schedules,
        "selected_schedule": selected_schedule,
        "faculty_history_results": faculty_history_results,
        "history_results": history_results,
        "faculty_count": len(faculty_history_results),
        "total_count": len(history_results),
    })

    return render(request, "admin/admin_past_evaluations.html", context)

def admin_logout(request):
    logout(request)
    return redirect("admin_login")



@role_required("UITC", "ADAA", "STAFF")
def admin_pending(request):
    schedules = EvaluationSchedule.objects.order_by('-start_datetime')
    selected_schedule = schedules.first()

    total_pending_evaluatees = 0
    total_done_evaluatees = 0
    department_tabs = []

    total_pending_heads = 0
    total_done_heads = 0
    total_heads = 0
    head_pending_rows = []
    head_done_rows = []

    if selected_schedule:
        # ==============================
        # HEAD MONITORING: ADAA -> HEADS
        # ==============================
        department_heads_for_adaa = (
            DepartmentHead.objects
            .filter(schedule=selected_schedule)
            .select_related("department")
            .order_by("department__name", "name")
        )

        for head in department_heads_for_adaa:
            total_heads += 1

            evaluation = (
                HeadEvaluation.objects
                .filter(
                    schedule=selected_schedule,
                    evaluatee_head=head
                )
                .order_by("-submitted_at", "-id")
                .first()
            )

            department_name = head.department.name if head.department else "No Department"
            department_code = head.department.code if head.department else "—"

            if evaluation and evaluation.status == "submitted":
                head_done_rows.append({
                    "head_name": evaluation.evaluatee_name or head.name,
                    "department": evaluation.evaluatee_department or department_name,
                    "department_code": department_code,
                    "status": "Done",
                    "evaluator_name": evaluation.evaluator_name or "ADAA",
                    "submitted_at": evaluation.submitted_at.strftime('%b %d, %Y %I:%M %p') if evaluation.submitted_at else "—",
                })
                total_done_heads += 1
            else:
                head_pending_rows.append({
                    "head_name": head.name,
                    "department": department_name,
                    "department_code": department_code,
                    "status": "Not Yet Finished",
                })
                total_pending_heads += 1

        # ===================================
        # FACULTY MONITORING: HEAD -> FACULTY
        # ===================================
        department_heads = (
            DepartmentHead.objects
            .filter(schedule=selected_schedule)
            .select_related('department')
            .order_by('department__name', 'name')
        )

        departments_map = {}

        for head in department_heads:
            dept = head.department
            dept_name = dept.name if dept else "No Department"
            dept_code = getattr(dept, 'code', '') if dept else ''

            if dept_name not in departments_map:
                departments_map[dept_name] = {
                    'name': dept_name,
                    'code': dept_code,
                    'slug': slugify(dept_name) or f"department-{len(departments_map)+1}",
                    'pending_rows': [],
                    'done_rows': [],
                    'pending_count': 0,
                    'done_count': 0,
                    'total_faculty': 0,
                }

            faculty_members = (
                FacultyMember.objects
                .filter(
                    schedule=selected_schedule,
                    department=dept
                )
                .select_related('department')
                .order_by('name')
            )

            for faculty in faculty_members:
                evaluation = (
                    FacultyEvaluation.objects
                    .filter(
                        schedule=selected_schedule,
                        evaluator_head=head,
                        evaluatee_faculty=faculty
                    )
                    .first()
                )

                if evaluation and evaluation.status == 'submitted':
                    departments_map[dept_name]['done_rows'].append({
                        'evaluatee_name': evaluation.evaluatee_name or faculty.name,
                        'evaluatee_department': evaluation.evaluatee_department or dept_name,
                        'status': 'Done',
                        'submitted_at': evaluation.submitted_at.strftime('%b %d, %Y %I:%M %p') if evaluation.submitted_at else '—',
                    })
                    departments_map[dept_name]['done_count'] += 1
                    total_done_evaluatees += 1
                else:
                    departments_map[dept_name]['pending_rows'].append({
                        'evaluatee_name': faculty.name,
                        'evaluatee_department': dept_name,
                        'status': 'Not Yet Finished',
                    })
                    departments_map[dept_name]['pending_count'] += 1
                    total_pending_evaluatees += 1

        for dept_name, dept_data in departments_map.items():
            dept_data['total_faculty'] = dept_data['pending_count'] + dept_data['done_count']

        department_tabs = sorted(departments_map.values(), key=lambda x: x['name'])

    context = {
        'active_page': 'pending',
        'selected_schedule': selected_schedule,
        'schedules': schedules,

        # Faculty pending data
        'department_tabs': department_tabs,
        'total_pending_evaluatees': total_pending_evaluatees,
        'total_done_evaluatees': total_done_evaluatees,

        # Head pending data
        'total_heads': total_heads,
        'total_pending_heads': total_pending_heads,
        'total_done_heads': total_done_heads,
        'head_pending_rows': head_pending_rows,
        'head_done_rows': head_done_rows,
    }

    return render(request, 'admin/admin_pending.html', context)






def _clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def _clean_header(value):
    if not value:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(value).strip().upper())


def _strip_rank_from_name(value):
    value = _clean_value(value)
    value = re.sub(r"^\s*\d+\s+", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _format_excel_number(value):
    if value is None or value == "":
        return ""

    try:
        number = float(value)

        if number.is_integer():
            return str(int(number))

        return f"{number:.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value).strip()


def _extract_number(value):
    if value is None:
        return None

    text = str(value).strip()

    if not text or text in ["-", "—"]:
        return None

    text = text.replace("%", "").replace(",", "")

    match = re.search(r"-?\d+(\.\d+)?", text)

    if not match:
        return None

    try:
        return float(match.group(0))
    except Exception:
        return None



def _format_percent_number(value):
    number = _extract_number(value)

    if number is None:
        return ""

    return f"{number:.2f}"


def _rating_to_percentage(computed_rating="", total_score="", mean_rating=""):
    computed = _extract_number(computed_rating)

    if computed is not None:
        if computed <= 5:
            computed = (computed / 5) * 100
        return f"{computed:.2f}"

    total = _extract_number(total_score)

    if total is not None and total > 0:
        return f"{(total / 75) * 100:.2f}"

    mean = _extract_number(mean_rating)

    if mean is not None:
        if mean <= 5:
            return f"{(mean / 5) * 100:.2f}"
        return f"{mean:.2f}"

    return ""

def _normalize_person_name(name):
    if not name:
        return ""

    name = str(name).strip().lower()
    name = re.sub(r"^\s*\d+\s+", "", name)

    # Convert "LAST, FIRST MIDDLE" into "FIRST MIDDLE LAST"
    if "," in name:
        parts = [p.strip() for p in name.split(",") if p.strip()]
        if len(parts) >= 2:
            name = f"{' '.join(parts[1:])} {parts[0]}"

    name = name.replace("-", " ")
    name = re.sub(r"[^a-zñ\s]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()

    ignored_tokens = {"jr", "sr", "ii", "iii", "iv", "v"}

    tokens = []

    for token in name.split():
        token = token.strip()

        if token in ignored_tokens:
            continue

        # Ignore middle initials only
        if len(token) == 1:
            continue

        tokens.append(token)

    # IMPORTANT:
    # Do NOT sort tokens.
    # Sorting makes "Juan Santos" and "Santos Juan" too loose.
    return " ".join(tokens)


def _get_cell(row, index):
    if index is None:
        return ""

    if index >= len(row):
        return ""

    return _clean_value(row[index].value)


def _find_header_row_and_indexes(ws):
    name_headers = {
        "NAME",
        "FULLNAME",
        "FACULTY",
        "FACULTYMEMBER",
        "FACULTYMEMBERNAME",
        "FACULTYNAME",
        "FACULTYEMPLOYEE",
        "FACULTYEMPLOYEENAME",
        "EVALUATEENAME",
        "EMPLOYEENAME",
        "TEACHER",
        "TEACHERNAME",
        "INSTRUCTOR",
        "INSTRUCTORNAME",
        "PERSONNAME",
    }

    first_name_headers = {
        "FIRSTNAME",
        "GIVENNAME",
    }

    middle_name_headers = {
        "MIDDLENAME",
        "MIDDLEINITIAL",
        "MI",
    }

    last_name_headers = {
        "LASTNAME",
        "SURNAME",
        "FAMILYNAME",
    }

    department_headers = {
        "DEPARTMENT",
        "DEPARTMENTCOLLEGE",
        "COLLEGE",
        "DEPT",
    }

    remarks_headers = {
        "REMARKS",
        "REMARK",
        "DESCRIPTION",
    }

    total_score_headers = {
        "AVERAGETOTALSCORE",
        "TOTALSCORE",
    }

    rating_priority = [
        "OVERALLAVERAGE",
        "MEAN",
        "AVERAGE",
        "AVERAGESCORE",
        "RATING",
        "OVERALLRATING",
        "TOTALRATING",
        "FINALSCORE",
        "FINALRATING",
        "SCORE",
    ]

    rank_headers = {
        "RANK",
        "FACULTYRANK",
        "ACADEMICRANK",
        "ACADEMICRANKPOSITION",
        "RANKPOSITION",
        "FACULTYRANKPOSITION",
        "CURRENTFACULTYRANK",
        "CURRENTFACULTYRANKPOSITION",
        "CURRENTACADEMICRANK",
        "CURRENTACADEMICRANKPOSITION",
        "POSITION",
        "POSITIONTITLE",
        "ACADEMICPOSITION",
        "ACADEMICPOSITIONTITLE",
        "DESIGNATION",
    }

    supervisor_headers = {
        "SUPERVISOR",
        "SUPERVISORS",
        "SUPERVISORNAME",
        "NAMEOFSUPERVISOR",
        "IMMEDIATESUPERVISOR",
        "EVALUATOR",
        "EVALUATORNAME",
        "RATER",
        "RATERNAME",
    }

    semester_headers = {
        "SEMESTER",
        "TERM",
        "SEMESTERTERM",
    }

    school_year_headers = {
        "SCHOOLYEAR",
        "ACADEMICYEAR",
        "AY",
        "SY",
    }

    period_headers = {
        "PERIOD",
        "EVALUATIONPERIOD",
        "SEMESTERTERMACADEMICYEAR",
        "SEMESTERANDACADEMICYEAR",
        "TERMACADEMICYEAR",
    }

    for row_number in range(1, min(ws.max_row, 15) + 1):
        row = list(ws[row_number])
        headers = [_clean_header(cell.value) for cell in row]

        name_index = None
        first_name_index = None
        middle_name_index = None
        last_name_index = None
        department_index = None
        rating_index = None
        computed_rating_index = None
        total_score_index = None
        remarks_index = None
        rank_index = None
        supervisor_index = None
        semester_index = None
        school_year_index = None
        period_index = None

        for index, header in enumerate(headers):
            if not header:
                continue

            if header in first_name_headers:
                first_name_index = index
                continue

            if header in middle_name_headers:
                middle_name_index = index
                continue

            if header in last_name_headers:
                last_name_index = index
                continue

            if header in supervisor_headers:
                supervisor_index = index
                continue

            if header in name_headers:
                name_index = index
                continue

            if header.endswith("NAME") and header not in supervisor_headers:
                name_index = index
                continue

            if header in department_headers:
                department_index = index
                continue

            if header in remarks_headers:
                remarks_index = index
                continue

            if header in rank_headers:
                rank_index = index
                continue

            if header in semester_headers:
                semester_index = index
                continue

            if header in school_year_headers:
                school_year_index = index
                continue

            if header in period_headers:
                period_index = index
                continue

            if header in total_score_headers:
                total_score_index = index
                continue

            if "COMPUTEDRATING" in header:
                computed_rating_index = index
                continue

        for priority_header in rating_priority:
            if priority_header in headers:
                rating_index = headers.index(priority_header)
                break

        if rating_index is None:
            for index, header in enumerate(headers):
                if (
                    "MEAN" in header
                    or "AVERAGE" in header
                    or "RATING" in header
                    or "SCORE" in header
                ):
                    if index != computed_rating_index:
                        rating_index = index
                        break

        if name_index is not None or (
            first_name_index is not None and last_name_index is not None
        ):
            return {
                "header_row": row_number,
                "name_index": name_index,
                "first_name_index": first_name_index,
                "middle_name_index": middle_name_index,
                "last_name_index": last_name_index,
                "department_index": department_index,
                "rating_index": rating_index,
                "computed_rating_index": computed_rating_index,
                "total_score_index": total_score_index,
                "remarks_index": remarks_index,
                "rank_index": rank_index,
                "supervisor_index": supervisor_index,
                "semester_index": semester_index,
                "school_year_index": school_year_index,
                "period_index": period_index,
            }

    return None

def _looks_like_pdf_table_row(row):
    values = [_clean_value(cell.value) for cell in row]

    if not values or not values[0]:
        return False

    first_cell = values[0]

    has_name = bool(re.search(r"[A-Za-zÑñ]{2,}", first_cell))
    starts_with_rank = bool(re.match(r"^\s*\d+\s+", first_cell))

    has_rating_next = False

    if len(row) > 1:
        has_rating_next = isinstance(row[1].value, (int, float))

    return has_name and (starts_with_rank or has_rating_next)

def _extract_people_from_excel(uploaded_file):
    records = []

    try:
        wb = load_workbook(uploaded_file, data_only=True)
    except Exception:
        raise ValueError("Invalid Excel file. Please upload a valid .xlsx file.")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        indexes = _find_header_row_and_indexes(ws)

        if indexes:
            start_row = indexes["header_row"] + 1

            for row_number in range(start_row, ws.max_row + 1):
                row = list(ws[row_number])

                name = ""

                if indexes["name_index"] is not None:
                    name = _get_cell(row, indexes["name_index"])

                if not name:
                    first_name = _get_cell(row, indexes["first_name_index"])
                    middle_name = _get_cell(row, indexes["middle_name_index"])
                    last_name = _get_cell(row, indexes["last_name_index"])

                    name = f"{first_name} {middle_name} {last_name}".strip()

                name = _strip_rank_from_name(name)

                if not name or name.lower() == "name":
                    continue

                normalized_name = _normalize_person_name(name)

                if not normalized_name:
                    continue

                records.append({
                    "name": name,
                    "normalized_name": normalized_name,
                    "department": _get_cell(row, indexes.get("department_index")),
                    "rank": _get_cell(row, indexes.get("rank_index")),
                    "supervisor": _get_cell(row, indexes.get("supervisor_index")),
                    "semester": _get_cell(row, indexes.get("semester_index")),
                    "school_year": _get_cell(row, indexes.get("school_year_index")),
                    "period": _get_cell(row, indexes.get("period_index")),
                    "rating": _format_excel_number(_get_cell(row, indexes.get("rating_index"))),
                    "computed_rating": _format_excel_number(_get_cell(row, indexes.get("computed_rating_index"))),
                    "total_score": _format_excel_number(_get_cell(row, indexes.get("total_score_index"))),
                    "remarks": _get_cell(row, indexes.get("remarks_index")),
                    "sheet": sheet_name,
                    "row": row_number,
                })

        else:
            for row_number in range(1, ws.max_row + 1):
                row = list(ws[row_number])

                if not _looks_like_pdf_table_row(row):
                    continue

                name = _strip_rank_from_name(row[0].value)
                normalized_name = _normalize_person_name(name)

                if not normalized_name:
                    continue

                # Get SET mean from the row.
                # Usually the SET exported table has:
                # Column 1 = Faculty name
                # Column 2 = Mean rating, example: 4.96
                set_mean_rating = ""

                for cell in row[1:]:
                    number = _extract_number(cell.value)

                    if number is None:
                        continue

                    # SET mean is normally 1.00 to 5.00.
                    # We take the first valid mean-like value after the name.
                    if 0 < number <= 5:
                        set_mean_rating = f"{number:.2f}"
                        break

                records.append({
                    "name": name,
                    "normalized_name": normalized_name,
                    "department": "",
                    "rank": "",
                    "supervisor": "",
                    "semester": "",
                    "school_year": "",
                    "period": "",

                    # This is the SET mean. _rating_to_percentage() will convert it:
                    # 4.96 -> 99.20
                    "rating": set_mean_rating,

                    "computed_rating": "",
                    "total_score": "",
                    "remarks": _get_cell(row, 2),
                    "sheet": sheet_name,
                    "row": row_number,
                })
    return records

def _token_set(normalized_name):
    return set(normalized_name.split())

def _get_first_last_from_normalized_name(normalized_name):
    tokens = [token for token in str(normalized_name or "").split() if token]

    if len(tokens) < 2:
        return "", ""

    first_name = tokens[0]
    last_name = tokens[-1]

    return first_name, last_name


def _is_safe_name_match(source_key, candidate_key):
    """
    Prevents matching different people just because they have the same surname.
    Example blocked:
    - Juan Santos
    - Maria Santos
    """

    source_first, source_last = _get_first_last_from_normalized_name(source_key)
    candidate_first, candidate_last = _get_first_last_from_normalized_name(candidate_key)

    if not source_first or not source_last or not candidate_first or not candidate_last:
        return False

    first_score = SequenceMatcher(None, source_first, candidate_first).ratio()
    last_score = SequenceMatcher(None, source_last, candidate_last).ratio()

    # First name must match or be almost identical
    if source_first != candidate_first and first_score < 0.92:
        return False

    # Last name must also be strict
    if source_last != candidate_last and last_score < 0.96:
        return False

    return True
def _find_fuzzy_match(source_key, candidate_keys, used_keys):
    source_tokens = _token_set(source_key)
    best_key = None
    best_score = 0

    for candidate_key in candidate_keys:
        if candidate_key in used_keys:
            continue

        candidate_tokens = _token_set(candidate_key)

        if not source_tokens or not candidate_tokens:
            continue

        # IMPORTANT:
        # Prevent same-surname-only matching.
        if not _is_safe_name_match(source_key, candidate_key):
            continue

        shorter_tokens = source_tokens
        longer_tokens = candidate_tokens

        if len(candidate_tokens) < len(source_tokens):
            shorter_tokens = candidate_tokens
            longer_tokens = source_tokens

        subset_match = shorter_tokens.issubset(longer_tokens) and len(shorter_tokens) >= 2
        similarity = SequenceMatcher(None, source_key, candidate_key).ratio()

        if subset_match:
            score = max(similarity, 0.90)
        elif similarity >= 0.88:
            score = similarity
        else:
            continue

        if score > best_score:
            best_score = score
            best_key = candidate_key

    return best_key, best_score

# ============================================================
# RECORDED SEF PRINT DATA
# ============================================================

SEF_PRINT_SECTION_TITLES = {
    "management_teaching_learning": "A. Management of Teaching and Learning",
    "content_knowledge_pedagogy_technology": "B. Content Knowledge, Pedagogy, and Technology",
    "commitment_transparency": "C. Commitment and Transparency",
}

SEF_PRINT_SECTION_ORDER = [
    "management_teaching_learning",
    "content_knowledge_pedagogy_technology",
    "commitment_transparency",
]

SEF_PRINT_CONTINUOUS_NUMBER_MAP = {
    "management_teaching_learning": {
        1: 1,
        2: 2,
        3: 3,
        4: 4,
        5: 5,
        6: 6,
    },
    "content_knowledge_pedagogy_technology": {
        1: 7,
        2: 8,
        3: 9,
        4: 10,
        5: 11,
    },
    "commitment_transparency": {
        1: 12,
        2: 13,
        3: 14,
        4: 15,
    },
}

SEF_PRINT_QUESTION_TEXTS = {
    1: "Comes to class on time.",
    2: "Submits updated syllabus, grade sheets, and other required reports on time.",
    3: "Maximizes the allocated time or learning hours effectively.",
    4: "Provides appropriate learning activities that facilitate students' critical thinking and creativity.",
    5: "Guides students to learn on their own, reflect on new ideas and experiences, and make decisions in accomplishing given tasks.",
    6: "Communicates constructive feedback to students for their academic growth.",
    7: "Demonstrates extensive and broad knowledge of the subject or course.",
    8: "Simplifies complex ideas in the lesson for ease of understanding.",
    9: "Integrates contemporary issues and developments in the discipline and/or daily life activities in the syllabus.",
    10: "Promotes active learning and student engagement by using appropriate teaching and learning resources including ICT tools and platforms.",
    11: "Uses appropriate assessments such as projects, exams, quizzes, and assignments aligned with the learning outcomes.",
    12: "Recognizes and values the unique diversity and individual differences among students.",
    13: "Assists students with their learning challenges during consultation hours.",
    14: "Provides immediate feedback on student outputs and performance.",
    15: "Provides transparent and clear criteria in rating student performance.",
}

SEF_PRINT_SUGGESTED_MEANS = {
    1: "Daily time record; Faculty schedule and timetable; Informal interview with students",
    2: "Documents submission log; Submission receipts or Acknowledgment Emails",
    3: "Class schedules and timetables; LMS logs; Informal interview with students",
    4: "Course syllabus; Learning plan; Classroom observation; Informal interview with students; LMS Logs",
    5: "Course syllabus; Learning plan; Student work samples; Class Observation; LMS Logs; Informal interview with students; Faculty Consultation Log",
    6: "Graded Student Work with Feedback; Faculty Consultation Log; Informal interview with students; Emails or Official correspondence; LMS Logs",
    7: "Course Syllabus; Learning Plan; IMs developed by the faculty; Informal interview with students; Mentorship or Thesis/Dissertation Advisory records",
    8: "Learning Plan; Course Syllabus; Classroom Observation; Informal Interview with students; Lecture notes and presentations; LMS Logs",
    9: "Course Syllabus; Learning Plan; Classroom Observation; Informal interview with students; LMS Logs; IMs developed by the faculty; Participation in Conferences, Webinars, and Training",
    10: "Course Syllabus; Learning Plan; Classroom Observation; Informal interview with students; LMS Logs; Multimedia Lecture Material; Student Work Samples",
    11: "Course Syllabus; Learning Plan; Informal interview with students; Assessment tools and rubrics; Exam and Quiz Samples; Graded Student Work Samples; LMS records",
    12: "Course Syllabus; Learning Plan; IMs developed by the faculty; Classroom Observation; Informal interview with students",
    13: "Course Syllabus; Faculty Consultation Log; Advisory Records; LMS Logs; Emails or Official Correspondence",
    14: "Graded Student Work Samples; Assessment tools and rubrics; Informal interview with students; LMS Logs; Emails or Official Correspondence; Faculty Consultation Log; Advising Reports",
    15: "Course Syllabus; Assessment Tools and Rubrics; Informal interview with students; LMS Records; Grade Sheets and Records",
}


def _sef_print_section_code(section_code, section_name=""):
    section_code = (section_code or "").strip()

    if section_code in SEF_PRINT_SECTION_TITLES:
        return section_code

    lowered = (section_name or "").strip().lower()

    if "management" in lowered and "teaching" in lowered:
        return "management_teaching_learning"

    if "content" in lowered or "pedagogy" in lowered or "technology" in lowered:
        return "content_knowledge_pedagogy_technology"

    if "commitment" in lowered or "transparency" in lowered:
        return "commitment_transparency"

    return section_code or "uncategorized"


def _sef_print_continuous_number(section_code, question_number):
    try:
        local_number = int(question_number or 0)
    except Exception:
        local_number = 0

    return SEF_PRINT_CONTINUOUS_NUMBER_MAP.get(section_code, {}).get(
        local_number,
        local_number,
    )


def _sef_print_rating(value):
    try:
        return f"{float(value or 0):.2f}"
    except (TypeError, ValueError):
        return "0.00"


def _sef_print_datetime(value):
    if not value:
        return ""

    try:
        return timezone.localtime(value).strftime("%B %d, %Y %I:%M %p")
    except Exception:
        return str(value)


def _build_sef_print_sections(responses):
    grouped = {section_code: [] for section_code in SEF_PRINT_SECTION_ORDER}

    for response in responses:
        section_code = _sef_print_section_code(
            response.section_code,
            response.section_name,
        )

        continuous_number = _sef_print_continuous_number(
            section_code,
            response.question_number,
        )

        question_text = (
            response.question_text
            or SEF_PRINT_QUESTION_TEXTS.get(continuous_number)
            or f"Benchmark Statement {continuous_number}"
        )

        grouped.setdefault(section_code, []).append({
            "number": continuous_number,
            "question_text": question_text,
            "suggested_means": SEF_PRINT_SUGGESTED_MEANS.get(continuous_number, "-"),
            "rating": _sef_print_rating(response.rating),
        })

    sections = []

    for section_code in SEF_PRINT_SECTION_ORDER:
        answers = sorted(
            grouped.get(section_code, []),
            key=lambda answer: answer["number"],
        )

        sections.append({
            "code": section_code,
            "title": SEF_PRINT_SECTION_TITLES[section_code],
            "answers": answers,
        })

    return sections

def _annex_rating_value(value):
    """
    Converts values like 5.00, 5, Decimal('5.00'), or '5.00'
    into clean Annex B rating values: 5, 4, 3, 2, 1.
    """
    if value is None:
        return ""

    try:
        number = float(value)

        if number.is_integer():
            return str(int(number))

        return f"{number:.2f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value).strip()


def _build_annex_rating_lookup(sections):
    """
    Creates:
    rating_1, rating_2, ..., rating_15

    This is needed for the exact Annex B design where ratings are shown
    under separate 5 / 4 / 3 / 2 / 1 cells.
    """
    rating_lookup = {
        f"rating_{number}": ""
        for number in range(1, 16)
    }

    for section in sections:
        for answer in section.get("answers", []):
            number = answer.get("number")
            rating = answer.get("rating")

            if number:
                rating_lookup[f"rating_{number}"] = _annex_rating_value(rating)

    return rating_lookup

def _get_available_sef_school_years():
    faculty_years = EvaluationSchedule.objects.filter(
        faculty_evaluations__status="submitted"
    ).values_list("academic_year", flat=True)

    head_years = EvaluationSchedule.objects.filter(
        head_evaluations__status="submitted"
    ).values_list("academic_year", flat=True)

    years = sorted(
        set(list(faculty_years) + list(head_years)),
        reverse=True
    )

    return years

def _get_faculty_rank_for_print(person):
    if not person:
        return ""

    possible_fields = ["academic_rank", "rank", "position", "faculty_rank"]

    for field in possible_fields:
        value = getattr(person, field, "")
        if value:
            return str(value).strip()

    return ""


def _get_print_extra_fields(person):
    if not person:
        return {
            "academic_rank": "",
            "course": "",
            "program_year": "",
        }

    return {
        "academic_rank": _get_faculty_rank_for_print(person),
        "course": str(getattr(person, "course", "") or "").strip(),
        "program_year": str(getattr(person, "program_year", "") or "").strip(),
    }
    
    
def _get_recorded_sef_results(selected_schedule_id="all"):
    grouped = {}

    selected_schedule_id = str(selected_schedule_id or "all").strip()
    now = timezone.localtime(timezone.now())

    def add_recorded_evaluation(
        *,
        evaluation,
        responses,
        target_name,
        target_department,
        target_rank="",
        target_course="",
        target_program_year="",
        target_type="Faculty",
        evaluator_name="",
        evaluator_department="",
        schedule=None,
        group_key="",
    ):
        if not target_name:
            return

        if not schedule:
            schedule = evaluation.schedule

        if group_key not in grouped:
            grouped[group_key] = {
                "_sort_start": schedule.start_datetime if schedule else timezone.now(),
                "print_id": "",
                "name": target_name,
                "department": target_department,
                "academic_rank": target_rank,
                "course": target_course,
                "program_year": target_program_year,
                "target_type": target_type,
                "schedule_id": schedule.id if schedule else "none",
                "academic_year": schedule.academic_year if schedule else "",
                "semester": schedule.semester if schedule else "",
                "title": schedule.title if schedule else "No Schedule",
                "schedule_label": (
                    f"{schedule.academic_year} | {schedule.semester} | {schedule.title}"
                    if schedule else "No Schedule"
                ),
                "supervisors": [],
                "supervisor_display": "",
                "evaluators": [],
            }

        responses = list(responses)
        section_values = defaultdict(list)

        for response in responses:
            section_code = _sef_print_section_code(
                response.section_code,
                response.section_name,
            )

            if section_code in SEF_PRINT_SECTION_TITLES:
                try:
                    section_values[section_code].append(float(response.rating or 0))
                except (TypeError, ValueError):
                    pass

        evaluator_sections = {}

        for section_code in SEF_PRINT_SECTION_ORDER:
            values = section_values.get(section_code, [])
            evaluator_sections[section_code] = _sef_print_rating(
                sum(values) / len(values) if values else 0
            )

        total_score = float(evaluation.total_score or 0)
        computed_rating = (total_score / 75) * 100 if total_score else 0

        sections = _build_sef_print_sections(responses)
        rating_lookup = _build_annex_rating_lookup(sections)

        evaluator_data = {
            "evaluation_id": evaluation.id,
            "evaluator_name": evaluator_name or "Unknown Supervisor",
            "evaluator_department": evaluator_department or "",
            "sections_summary": evaluator_sections,
            "average_score": _sef_print_rating(evaluation.average_score),
            "total_score": _sef_print_rating(total_score),
            "computed_rating": _sef_print_rating(computed_rating),
            "comments": evaluation.comments or "",
            "submitted_at": _sef_print_datetime(evaluation.submitted_at),
            "sections": sections,
        }

        evaluator_data.update(rating_lookup)

        grouped[group_key]["evaluators"].append(evaluator_data)

    # =========================
    # FACULTY SEF RECORDS
    # Head evaluates faculty
    # =========================
    faculty_evaluations = (
        FacultyEvaluation.objects
        .filter(status="submitted")
        .select_related(
            "schedule",
            "evaluatee_faculty__department",
            "evaluator_head__department",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(FacultyEvaluationResponse),
            )
        )
        .order_by(
            "-schedule__start_datetime",
            "evaluatee_name",
            "evaluator_name",
            "submitted_at",
        )
    )

    if selected_schedule_id and selected_schedule_id != "all":
        if not selected_schedule_id.isdigit():
            return []

        faculty_evaluations = faculty_evaluations.filter(
            schedule_id=selected_schedule_id
        )
    else:
        faculty_evaluations = faculty_evaluations.filter(
            schedule__end_datetime__lt=now
        )

    for evaluation in faculty_evaluations:
        faculty_obj = evaluation.evaluatee_faculty

        target_name = (
            evaluation.evaluatee_name
            or (faculty_obj.name if faculty_obj else "Unknown Faculty")
        )

        target_department = (
            evaluation.evaluatee_department
            or (
                faculty_obj.department.name
                if faculty_obj and faculty_obj.department
                else ""
            )
        )

        supervisor_name = (
            evaluation.evaluator_name
            or (
                evaluation.evaluator_head.name
                if evaluation.evaluator_head
                else ""
            )
        )

        supervisor_department = (
            evaluation.evaluator_department
            or (
                evaluation.evaluator_head.department.name
                if evaluation.evaluator_head and evaluation.evaluator_head.department
                else ""
            )
        )

        schedule = evaluation.schedule
        faculty_key = evaluation.evaluatee_faculty_id or _normalize_person_name(target_name)

        faculty_extra = _get_print_extra_fields(faculty_obj)

        add_recorded_evaluation(
            evaluation=evaluation,
            responses=evaluation.responses.all(),
            target_name=target_name,
            target_department=target_department,
            target_rank=faculty_extra.get("academic_rank", ""),
            target_course=faculty_extra.get("course", ""),
            target_program_year=faculty_extra.get("program_year", ""),
            target_type="Faculty",
            evaluator_name=supervisor_name,
            evaluator_department=supervisor_department,
            schedule=schedule,
            group_key=f"faculty-{schedule.id if schedule else 'none'}-{faculty_key}",
        )

    # =========================
    # HEAD SEF RECORDS
    # ADAA evaluates department heads
    # =========================
    head_evaluations = (
        HeadEvaluation.objects
        .filter(status="submitted")
        .select_related(
            "schedule",
            "evaluatee_head__department",
            "evaluator_officer",
        )
        .prefetch_related(
            Prefetch(
                "responses",
                queryset=_ordered_response_queryset(HeadEvaluationResponse),
            )
        )
        .order_by(
            "-schedule__start_datetime",
            "evaluatee_name",
            "evaluator_name",
            "submitted_at",
        )
    )

    if selected_schedule_id and selected_schedule_id != "all":
        if not selected_schedule_id.isdigit():
            return []

        head_evaluations = head_evaluations.filter(
            schedule_id=selected_schedule_id
        )
    else:
        head_evaluations = head_evaluations.filter(
            schedule__end_datetime__lt=now
        )

    for evaluation in head_evaluations:
        head_obj = evaluation.evaluatee_head

        target_name = (
            evaluation.evaluatee_name
            or (head_obj.name if head_obj else "Unknown Department Head")
        )

        target_department = (
            evaluation.evaluatee_department
            or (
                head_obj.department.name
                if head_obj and head_obj.department
                else ""
            )
        )

        supervisor_name = (
            evaluation.evaluator_name
            or (
                evaluation.evaluator_officer.name
                if evaluation.evaluator_officer
                else ""
            )
        )

        supervisor_department = "Office of the ADAA"

        schedule = evaluation.schedule
        head_key = evaluation.evaluatee_head_id or _normalize_person_name(target_name)

        head_extra = _get_print_extra_fields(head_obj)

        add_recorded_evaluation(
            evaluation=evaluation,
            responses=evaluation.responses.all(),
            target_name=target_name,
            target_department=target_department,
            target_rank=head_extra.get("academic_rank", ""),
            target_course=head_extra.get("course", ""),
            target_program_year=head_extra.get("program_year", ""),
            target_type="Department Head",
            evaluator_name=supervisor_name,
            evaluator_department=supervisor_department,
            schedule=schedule,
            group_key=f"head-{schedule.id if schedule else 'none'}-{head_key}",
        )

    recorded_results = list(grouped.values())

    recorded_results.sort(
        key=lambda item: (
            item.get("_sort_start") or timezone.now(),
            item.get("name", "").lower(),
        ),
        reverse=True,
    )

    for index, item in enumerate(recorded_results, start=1):
        item.pop("_sort_start", None)
        item["print_id"] = f"recorded-{index}"
        item["evaluator_count"] = len(item.get("evaluators", []))

        seen_supervisors = set()
        supervisors = []

        for evaluator in item.get("evaluators", []):
            supervisor_name = evaluator.get("evaluator_name", "").strip()
            supervisor_department = evaluator.get("evaluator_department", "").strip()

            if not supervisor_name:
                continue

            supervisor_key = supervisor_name.lower()

            if supervisor_key in seen_supervisors:
                continue

            seen_supervisors.add(supervisor_key)

            supervisors.append({
                "name": supervisor_name,
                "department": supervisor_department,
            })

        item["supervisors"] = supervisors
        item["supervisor_display"] = "; ".join(
            supervisor["name"] for supervisor in supervisors
        )

    return recorded_results


def _get_recorded_sef_records_for_matching(selected_schedule_id="all"):
    """
    Build SEF matching records from submitted SEF database records.
    This replaces the old SEF Excel upload in the SEF + SET workflow.
    """

    recorded_results = _get_recorded_sef_results(selected_schedule_id)
    sef_records = []

    for item in recorded_results:
        name = (item.get("name") or "").strip()
        normalized_name = _normalize_person_name(name)

        if not name or not normalized_name:
            continue

        computed_values = []
        total_values = []

        for evaluator in item.get("evaluators", []):
            computed_value = _extract_number(evaluator.get("computed_rating", ""))
            total_value = _extract_number(evaluator.get("total_score", ""))

            if computed_value is not None:
                computed_values.append(computed_value)

            if total_value is not None:
                total_values.append(total_value)

        average_computed = (
            sum(computed_values) / len(computed_values)
            if computed_values
            else None
        )

        average_total = (
            sum(total_values) / len(total_values)
            if total_values
            else None
        )

        sef_records.append({
            "name": name,
            "normalized_name": normalized_name,
            "department": item.get("department", ""),
            "rank": item.get("academic_rank", "") or item.get("rank", ""),
            "supervisor": item.get("supervisor_display", ""),
            "semester": item.get("semester", ""),
            "school_year": item.get("academic_year", ""),
            "period": item.get("schedule_label", ""),
            "rating": "",
            "computed_rating": f"{average_computed:.2f}" if average_computed is not None else "",
            "total_score": f"{average_total:.2f}" if average_total is not None else "",
            "remarks": "",
            "sheet": "Recorded SEF",
            "row": "",
        })

    return sef_records




def _get_uploaded_person_extra_for_print(name):
    """
    Gets academic rank, course, program year, and department from uploaded
    FacultyMember or DepartmentHead records using the person's name.
    Used as fallback when SEF/SET Excel files do not contain rank.
    """
    empty_data = {
        "academic_rank": "",
        "course": "",
        "program_year": "",
        "department": "",
    }

    normalized_target = _normalize_person_name(name)

    if not normalized_target:
        return empty_data

    schedule = _get_latest_schedule_with_uploaded_data()

    faculty_qs = FacultyMember.objects.select_related("department").all()
    head_qs = DepartmentHead.objects.select_related("department").all()

    if schedule:
        faculty_qs = faculty_qs.filter(schedule=schedule)
        head_qs = head_qs.filter(schedule=schedule)

    for faculty in faculty_qs:
        if _normalize_person_name(faculty.name) == normalized_target:
            return {
                "academic_rank": str(faculty.academic_rank or "").strip(),
                "course": str(faculty.course or "").strip(),
                "program_year": str(faculty.program_year or "").strip(),
                "department": faculty.department.name if faculty.department else "",
            }

    for head in head_qs:
        if _normalize_person_name(head.name) == normalized_target:
            return {
                "academic_rank": str(head.academic_rank or "").strip(),
                "course": str(head.course or "").strip(),
                "program_year": str(head.program_year or "").strip(),
                "department": head.department.name if head.department else "",
            }

    return empty_data

def _build_sef_set_match_results(
    sef_records,
    set_records,
    sef_filename="",
    set_filename="",
    rank_records=None,
):
    sef_map = {}
    set_map = {}
    rank_lookup = _build_rank_lookup_from_records(rank_records or [])

    for record in sef_records:
        normalized_name = (
            record.get("normalized_name")
            or _normalize_person_name(record.get("name", ""))
        )

        if not normalized_name:
            continue

        record["normalized_name"] = normalized_name
        sef_map.setdefault(normalized_name, []).append(record)

    for record in set_records:
        normalized_name = (
            record.get("normalized_name")
            or _normalize_person_name(record.get("name", ""))
        )

        if not normalized_name:
            continue

        record["normalized_name"] = normalized_name
        set_map.setdefault(normalized_name, []).append(record)

    exact_keys = sorted(set(sef_map.keys()) & set(set_map.keys()))

    matched_results = []
    used_set_keys = set()
    rank_matched_count = 0

    def build_item(sef_record, set_record, match_type, match_score):
        nonlocal rank_matched_count

        sef_name = sef_record.get("name", "")
        set_name = set_record.get("name", "")

        normalized_name = (
            sef_record.get("normalized_name")
            or set_record.get("normalized_name")
            or _normalize_person_name(sef_name)
            or _normalize_person_name(set_name)
        )

        rank_data = _find_rank_data_for_person(
            [
                sef_name,
                set_name,
                sef_record.get("normalized_name", ""),
                set_record.get("normalized_name", ""),
                normalized_name,
            ],
            rank_lookup,
        ) or {}

        fallback_extra = _get_uploaded_person_extra_for_print(sef_name or set_name)

        department_value = (
            sef_record.get("department", "")
            or set_record.get("department", "")
            or fallback_extra.get("department", "")
            or rank_data.get("department", "")
        )

        rank_value = (
            rank_data.get("rank", "")
            or sef_record.get("rank", "")
            or set_record.get("rank", "")
            or fallback_extra.get("academic_rank", "")
        )

        if rank_value:
            rank_matched_count += 1

        sef_percentage = (sef_record.get("computed_rating") or "").strip()

        if not sef_percentage:
            sef_percentage = _rating_to_percentage(
                computed_rating=sef_record.get("computed_rating", ""),
                total_score=sef_record.get("total_score", ""),
                mean_rating=sef_record.get("rating", ""),
            )

        set_percentage = _rating_to_percentage(
            computed_rating=set_record.get("computed_rating", ""),
            total_score=set_record.get("total_score", ""),
            mean_rating=set_record.get("rating", ""),
        )

        return {
            "name": sef_name or set_name,
            "normalized_name": normalized_name,
            "department": department_value,
            "rank": rank_value,
            "supervisor": sef_record.get("supervisor", "") or set_record.get("supervisor", ""),
            "semester": sef_record.get("semester", "") or set_record.get("semester", ""),
            "school_year": sef_record.get("school_year", "") or set_record.get("school_year", ""),
            "period": sef_record.get("period", "") or set_record.get("period", ""),

            "sef_name": sef_name,
            "set_name": set_name,

            "sef_rating": sef_percentage,
            "sef_computed_rating": sef_percentage,
            "sef_total_score": sef_record.get("total_score", ""),

            "set_rating": set_percentage,
            "set_remarks": set_record.get("remarks", ""),

            "sef_sheet": sef_record.get("sheet", ""),
            "set_sheet": set_record.get("sheet", ""),

            "match_type": match_type,
            "match_score": match_score,
        }

    for key in exact_keys:
        sef_record = sef_map[key][0]
        set_record = set_map[key][0]
        used_set_keys.add(key)

        matched_results.append(
            build_item(
                sef_record=sef_record,
                set_record=set_record,
                match_type="Exact",
                match_score="100%",
            )
        )

    remaining_sef_keys = [
        key for key in sef_map.keys()
        if key not in exact_keys
    ]

    remaining_set_keys = [
        key for key in set_map.keys()
        if key not in exact_keys
    ]

    fuzzy_matched_sef_keys = set()

    for sef_key in remaining_sef_keys:
        matched_set_key, score = _find_fuzzy_match(
            sef_key,
            remaining_set_keys,
            used_set_keys,
        )

        if matched_set_key:
            sef_record = sef_map[sef_key][0]
            set_record = set_map[matched_set_key][0]

            fuzzy_matched_sef_keys.add(sef_key)
            used_set_keys.add(matched_set_key)

            matched_results.append(
                build_item(
                    sef_record=sef_record,
                    set_record=set_record,
                    match_type="Possible Match",
                    match_score=f"{round(score * 100)}%",
                )
            )

    matched_results.sort(
        key=lambda item: str(item.get("name", "")).lower()
    )

    unmatched_sef = []
    unmatched_set = []

    def build_unmatched_sef_item(sef_record):
        sef_name = sef_record.get("name", "")
        normalized_name = (
            sef_record.get("normalized_name")
            or _normalize_person_name(sef_name)
        )

        rank_data = _find_rank_data_for_person(
            [
                sef_name,
                normalized_name,
            ],
            rank_lookup,
        ) or {}

        fallback_extra = _get_uploaded_person_extra_for_print(sef_name)

        department_value = (
            sef_record.get("department", "")
            or fallback_extra.get("department", "")
            or rank_data.get("department", "")
        )

        rank_value = (
            rank_data.get("rank", "")
            or sef_record.get("rank", "")
            or fallback_extra.get("academic_rank", "")
        )

        sef_percentage = (sef_record.get("computed_rating") or "").strip()

        if not sef_percentage:
            sef_percentage = _rating_to_percentage(
                computed_rating=sef_record.get("computed_rating", ""),
                total_score=sef_record.get("total_score", ""),
                mean_rating=sef_record.get("rating", ""),
            )

        return {
            "name": sef_name,
            "normalized_name": normalized_name,
            "department": department_value,
            "rank": rank_value,
            "supervisor": sef_record.get("supervisor", ""),
            "semester": sef_record.get("semester", ""),
            "school_year": sef_record.get("school_year", ""),
            "period": sef_record.get("period", ""),
            "rating": sef_record.get("rating", ""),
            "computed_rating": sef_percentage,
            "sef_computed_rating": sef_percentage,
            "total_score": sef_record.get("total_score", ""),
            "sheet": sef_record.get("sheet", ""),
            "row": sef_record.get("row", ""),
            "match_type": "SEF Only",
            "match_score": "",
        }

    for key, records in sef_map.items():
        if key not in exact_keys and key not in fuzzy_matched_sef_keys:
            unmatched_sef.append(build_unmatched_sef_item(records[0]))

    for key, records in set_map.items():
        if key not in used_set_keys:
            unmatched_set.append(records[0])

    unmatched_sef.sort(
        key=lambda item: str(item.get("name", "")).lower()
    )

    unmatched_set.sort(
        key=lambda item: str(item.get("name", "")).lower()
    )

    return {
        "sef_filename": sef_filename,
        "set_filename": set_filename,
        "total_sef": len(sef_records),
        "total_set": len(set_records),
        "total_matched": len(matched_results),
        "total_rank_matched": rank_matched_count,
        "total_unmatched_sef": len(unmatched_sef),
        "total_unmatched_set": len(unmatched_set),
        "matched_results": matched_results,
        "unmatched_sef": unmatched_sef,
        "unmatched_set": unmatched_set,
        "generated_at": timezone.localtime(timezone.now()).strftime("%B %d, %Y %I:%M %p"),
    }
    
    
def _save_sef_set_results_to_db(request, sef_file=None, set_file=None, match_results=None):
    match_results = match_results or {}

    try:
        if sef_file:
            sef_file.seek(0)

        if set_file:
            set_file.seek(0)
    except Exception:
        pass


    if request.user.is_authenticated:
        SEFSETUploadBatch.objects.filter(
            uploaded_by=request.user,
            is_active=True
        ).update(is_active=False)
        
    batch = SEFSETUploadBatch.objects.create(
        sef_file=sef_file if sef_file else "",
        set_file=set_file if set_file else "",
        sef_filename=match_results.get("sef_filename", "Recorded SEF from Database"),
        set_filename=(set_file.name if set_file else match_results.get("set_filename", "")),
        total_sef=match_results.get("total_sef", 0),
        total_set=match_results.get("total_set", 0),
        total_matched=match_results.get("total_matched", 0),
        total_unmatched_sef=match_results.get("total_unmatched_sef", 0),
        total_unmatched_set=match_results.get("total_unmatched_set", 0),
        unmatched_sef_json=match_results.get("unmatched_sef", []),
        unmatched_set_json=match_results.get("unmatched_set", []),
        uploaded_by=request.user if request.user.is_authenticated else None,
    )

    matched_objects = []

    for item in match_results.get("matched_results", []):
        matched_objects.append(
            SEFSETMatchedResult(
                batch=batch,
                name=item.get("name", ""),
                normalized_name=item.get("normalized_name", ""),
                department=item.get("department", ""),
                rank=item.get("rank", ""),
                supervisor=item.get("supervisor", ""),
                semester=item.get("semester", ""),
                school_year=item.get("school_year", ""),
                period=item.get("period", ""),
                sef_name=item.get("sef_name", ""),
                set_name=item.get("set_name", ""),
                sef_rating=item.get("sef_rating", ""),
                sef_computed_rating=item.get("sef_computed_rating", ""),
                sef_total_score=item.get("sef_total_score", ""),
                set_rating=item.get("set_rating", ""),
                set_remarks=item.get("set_remarks", ""),
                sef_sheet=item.get("sef_sheet", ""),
                set_sheet=item.get("set_sheet", ""),
                match_type=item.get("match_type", ""),
                match_score=item.get("match_score", ""),
            )
        )

    SEFSETMatchedResult.objects.bulk_create(matched_objects)

    return batch


def _load_sef_set_results_from_db(batch):
    if not batch:
        return None

    matched_results = []

    for item in batch.matched_results.all().order_by("name"):
        matched_results.append({
            "name": item.name,
            "normalized_name": item.normalized_name,
            "department": item.department,
            "rank": item.rank,
            "supervisor": item.supervisor,
            "semester": item.semester,
            "school_year": item.school_year,
            "period": item.period,
            "sef_name": item.sef_name,
            "set_name": item.set_name,
            "sef_rating": item.sef_rating,
            "sef_computed_rating": item.sef_computed_rating,
            "sef_total_score": item.sef_total_score,
            "set_rating": item.set_rating,
            "set_remarks": item.set_remarks,
            "sef_sheet": item.sef_sheet,
            "set_sheet": item.set_sheet,
            "match_type": item.match_type,
            "match_score": item.match_score,
        })

    unmatched_sef = batch.unmatched_sef_json or []
    unmatched_set = batch.unmatched_set_json or []

    unmatched_sef = sorted(
        unmatched_sef,
        key=lambda item: str(item.get("name", "")).lower()
    )

    unmatched_set = sorted(
        unmatched_set,
        key=lambda item: str(item.get("name", "")).lower()
    )

    return {
        "batch_id": batch.id,
        "sef_filename": batch.sef_filename,
        "set_filename": batch.set_filename,
        "total_sef": batch.total_sef,
        "total_set": batch.total_set,
        "total_matched": batch.total_matched,
        "total_unmatched_sef": batch.total_unmatched_sef,
        "total_unmatched_set": batch.total_unmatched_set,
        "matched_results": matched_results,
        "unmatched_sef": unmatched_sef,
        "unmatched_set": unmatched_set,
        "generated_at": timezone.localtime(batch.generated_at).strftime("%B %d, %Y %I:%M %p"),
    }
    
def _get_available_recorded_sef_schedules():
    now = timezone.localtime(timezone.now())

    faculty_schedule_ids = FacultyEvaluation.objects.filter(
        status="submitted",
        schedule__end_datetime__lt=now,
    ).values_list("schedule_id", flat=True)

    head_schedule_ids = HeadEvaluation.objects.filter(
        status="submitted",
        schedule__end_datetime__lt=now,
    ).values_list("schedule_id", flat=True)

    schedule_ids = set(
        list(faculty_schedule_ids) + list(head_schedule_ids)
    )

    schedule_ids.discard(None)

    return (
        EvaluationSchedule.objects
        .filter(id__in=schedule_ids)
        .order_by("-start_datetime", "-created_at")
    )


@role_required("UITC", "ADAA", "STAFF")
def admin_sef_set(request):
    page_mode = request.GET.get("mode", "home").strip().lower()

    if page_mode not in ["home", "recorded", "excel"]:
        page_mode = "home"

    # ======================================================
    # CLEAR ACTIVE SET UPLOAD + MATCHED RECORDED SEF
    # ======================================================
    if request.GET.get("clear") == "1":
        request.session.pop("sef_set_upload_batch_id", None)

        SEFSETUploadBatch.objects.filter(
            uploaded_by=request.user,
            is_active=True
        ).update(is_active=False)

        messages.success(
            request,
            "Current SET upload and matched SEF records cleared from the active printable list."
        )
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    # ======================================================
    # RECORDED SEF SCHEDULES
    # ======================================================
    recorded_sef_schedules = _get_available_recorded_sef_schedules()

    selected_recorded_schedule = request.GET.get("schedule", "all").strip()

    valid_schedule_ids = [
        str(schedule.id)
        for schedule in recorded_sef_schedules
    ]

    if selected_recorded_schedule != "all" and selected_recorded_schedule not in valid_schedule_ids:
        selected_recorded_schedule = "all"

    # ======================================================
    # EXCEL MODE: RECORDED SEF + UPLOADED SET
    # ======================================================
    results = None

    if page_mode == "excel":
        batch = None
        batch_id = request.session.get("sef_set_upload_batch_id")

        if batch_id:
            batch = SEFSETUploadBatch.objects.filter(
                id=batch_id,
                is_active=True
            ).first()

        if not batch:
            batch = (
                SEFSETUploadBatch.objects
                .filter(is_active=True, uploaded_by=request.user)
                .order_by("-generated_at", "-id")
                .first()
            )

        results = _load_sef_set_results_from_db(batch)

    # ======================================================
    # RECORDED MODE: ANNEX B SEF ONLY
    # ======================================================
    recorded_sef_results = []

    if page_mode == "recorded":
        recorded_sef_results = _get_recorded_sef_results(selected_recorded_schedule)

    # Old compatibility variables. Keep them so old template references will not crash.
    sef_school_years = sorted(
        {
            schedule.academic_year
            for schedule in recorded_sef_schedules
            if schedule.academic_year
        },
        reverse=True
    )

    selected_school_year = request.GET.get("school_year", "all").strip()

    if selected_school_year != "all" and selected_school_year not in sef_school_years:
        selected_school_year = "all"

    return render(request, "admin/admin_sef_set.html", {
        "active_page": "sef_set",
        "page_mode": page_mode,

        "results": results,

        "recorded_sef_results": recorded_sef_results,
        "recorded_sef_schedules": recorded_sef_schedules,
        "selected_recorded_schedule": selected_recorded_schedule,

        "sef_school_years": sef_school_years,
        "selected_school_year": selected_school_year,
    })
    
@role_required("UITC", "ADAA", "STAFF")
def excel_matcher(request):
    if request.method != "POST":
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    set_file = request.FILES.get("set_file")
    rank_file = request.FILES.get("rank_file")
    selected_schedule_id = (request.POST.get("schedule_id") or "all").strip()

    if not set_file:
        messages.error(request, "Please upload the SET Excel file.")
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    if not set_file.name.lower().endswith(".xlsx"):
        messages.error(request, "Please upload the SET file in .xlsx format.")
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    if rank_file and not rank_file.name.lower().endswith(".xlsx"):
        messages.error(request, "Please upload the rank file in .xlsx format.")
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    recorded_sef_schedules = _get_available_recorded_sef_schedules()

    valid_schedule_ids = [
        str(schedule.id)
        for schedule in recorded_sef_schedules
    ]

    if selected_schedule_id != "all" and selected_schedule_id not in valid_schedule_ids:
        messages.error(
            request,
            "The selected SEF schedule is invalid or has no submitted SEF records."
        )
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    try:
        sef_records = _get_recorded_sef_records_for_matching(selected_schedule_id)
        set_records = _extract_people_from_excel(set_file)
        rank_records = _extract_people_from_excel(rank_file) if rank_file else []

    except ValueError as e:
        messages.error(request, str(e))
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")
    except Exception as e:
        messages.error(request, f"SET processing failed: {str(e)}")
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    if not sef_records:
        messages.error(
            request,
            "No submitted SEF records were found for the selected schedule."
        )
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    if not set_records:
        messages.error(
            request,
            "No valid names were found in the SET file."
        )
        return redirect(f"{reverse('admin_sef_set')}?mode=excel")

    schedule_label = "All Past Recorded SEF"

    if selected_schedule_id != "all":
        selected_schedule = recorded_sef_schedules.filter(
            id=selected_schedule_id
        ).first()

        if selected_schedule:
            schedule_label = (
                f"{selected_schedule.academic_year} | "
                f"{selected_schedule.semester} | "
                f"{selected_schedule.title}"
            )

    match_results = _build_sef_set_match_results(
        sef_records=sef_records,
        set_records=set_records,
        sef_filename=f"Recorded SEF - {schedule_label}",
        set_filename=set_file.name,
        rank_records=rank_records,
    )

    batch = _save_sef_set_results_to_db(
        request=request,
        sef_file=None,
        set_file=set_file,
        match_results=match_results,
    )

    request.session["sef_set_upload_batch_id"] = batch.id
    rank_message = ""

    if rank_file:
        rank_message = (
            f" Rank file was also processed. "
            f"{len(rank_records)} rank row(s) were read from the uploaded rank file. "
            f"{match_results.get('total_rank_matched', 0)} rank(s) were attached to matched Annex D records."
        )
        
    messages.success(
        request,
        f"Matching complete using recorded SEF records. "
        f"{match_results['total_matched']} matched name(s), "
        f"{match_results['total_unmatched_sef']} SEF-only name(s), and "
        f"{match_results['total_unmatched_set']} SET-only name(s) found."
        f"{rank_message}"
    )

    return redirect(f"{reverse('admin_sef_set')}?mode=excel")

@role_required("UITC", "ADAA")
def admin_user_management(request):
    current_role = get_admin_role(request.user)

    if request.method == "POST":
        action = request.POST.get("action")

        # =========================
        # CREATE USER
        # =========================
        if action == "create_user":
            username = (request.POST.get("username") or "").strip()
            email = (request.POST.get("email") or "").strip()
            first_name = (request.POST.get("first_name") or "").strip()
            last_name = (request.POST.get("last_name") or "").strip()
            role = (request.POST.get("role") or "").strip()
            password1 = request.POST.get("password1") or ""
            password2 = request.POST.get("password2") or ""

            if current_role == "ADAA" and role != "STAFF":
                messages.error(request, "ADAA can only create STAFF accounts.")
                return redirect("admin_user_management")

            if current_role == "UITC" and role not in ["ADAA", "STAFF"]:
                messages.error(request, "Please select a valid role.")
                return redirect("admin_user_management")

            if not username or not email or not role or not password1 or not password2:
                messages.error(request, "Please complete all required fields.")
                return redirect("admin_user_management")

            if password1 != password2:
                messages.error(request, "Passwords do not match.")
                return redirect("admin_user_management")

            if len(password1) < 8:
                messages.error(request, "Password must be at least 8 characters long.")
                return redirect("admin_user_management")

            if User.objects.filter(username__iexact=username).exists():
                messages.error(request, "Username already exists.")
                return redirect("admin_user_management")

            if User.objects.filter(email__iexact=email).exists():
                messages.error(request, "Email already exists.")
                return redirect("admin_user_management")

            try:
                group = Group.objects.get(name=role)
            except Group.DoesNotExist:
                messages.error(request, "Selected role does not exist. Please create the group first.")
                return redirect("admin_user_management")

            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name,
                is_staff=True,
                is_active=True,
                is_superuser=False,
            )

            user.groups.add(group)

            messages.success(request, f"{role} account created successfully.")
            return redirect("admin_user_management")

        # =========================
        # RESET PASSWORD
        # =========================
        elif action == "reset_password":
            user_id = request.POST.get("user_id")
            password1 = request.POST.get("password1") or ""
            password2 = request.POST.get("password2") or ""

            target_user = get_object_or_404(User, id=user_id, is_staff=True)
            target_role = get_admin_role(target_user)

            if target_user.id == request.user.id:
                messages.error(request, "You cannot reset your own password here.")
                return redirect("admin_user_management")

            if target_user.is_superuser:
                messages.error(request, "You cannot reset a UITC superuser password here.")
                return redirect("admin_user_management")

            if current_role == "ADAA" and target_role != "STAFF":
                messages.error(request, "ADAA can only reset STAFF passwords.")
                return redirect("admin_user_management")

            if current_role == "UITC" and target_role not in ["ADAA", "STAFF"]:
                messages.error(request, "You can only reset ADAA or STAFF accounts here.")
                return redirect("admin_user_management")

            if not password1 or not password2:
                messages.error(request, "Please enter and confirm the new password.")
                return redirect("admin_user_management")

            if password1 != password2:
                messages.error(request, "Passwords do not match.")
                return redirect("admin_user_management")

            if len(password1) < 8:
                messages.error(request, "Password must be at least 8 characters long.")
                return redirect("admin_user_management")

            target_user.set_password(password1)
            target_user.save()

            messages.success(request, f"Password reset successfully for {target_user.username}.")
            return redirect("admin_user_management")

        # =========================
        # ACTIVATE / DEACTIVATE USER
        # =========================
        elif action == "toggle_active":
            user_id = request.POST.get("user_id")

            target_user = get_object_or_404(User, id=user_id, is_staff=True)
            target_role = get_admin_role(target_user)

            if target_user.id == request.user.id:
                messages.error(request, "You cannot deactivate your own account.")
                return redirect("admin_user_management")

            if target_user.is_superuser:
                messages.error(request, "You cannot deactivate a UITC superuser account here.")
                return redirect("admin_user_management")

            if current_role == "ADAA" and target_role != "STAFF":
                messages.error(request, "ADAA can only activate or deactivate STAFF accounts.")
                return redirect("admin_user_management")

            if current_role == "UITC" and target_role not in ["ADAA", "STAFF"]:
                messages.error(request, "You can only activate or deactivate ADAA or STAFF accounts here.")
                return redirect("admin_user_management")

            target_user.is_active = not target_user.is_active
            target_user.save()

            status = "activated" if target_user.is_active else "deactivated"
            messages.success(request, f"{target_user.username} has been {status}.")
            return redirect("admin_user_management")

        # =========================
        # DELETE USER
        # =========================
        elif action == "delete_user":
            user_id = request.POST.get("user_id")

            target_user = get_object_or_404(User, id=user_id, is_staff=True)
            target_role = get_admin_role(target_user)

            if target_user.id == request.user.id:
                messages.error(request, "You cannot delete your own account.")
                return redirect("admin_user_management")

            if target_user.is_superuser:
                messages.error(request, "You cannot delete a UITC superuser account here.")
                return redirect("admin_user_management")

            if current_role == "ADAA" and target_role != "STAFF":
                messages.error(request, "ADAA can only delete STAFF accounts.")
                return redirect("admin_user_management")

            if current_role == "UITC" and target_role not in ["ADAA", "STAFF"]:
                messages.error(request, "UITC can only delete ADAA or STAFF accounts here.")
                return redirect("admin_user_management")

            deleted_username = target_user.username
            target_user.delete()

            messages.success(request, f"Account '{deleted_username}' has been deleted successfully.")
            return redirect("admin_user_management")

        else:
            messages.error(request, "Invalid action.")
            return redirect("admin_user_management")

    # =========================
    # DISPLAY USERS
    # =========================
    if current_role == "UITC":
        users = (
            User.objects
            .filter(is_staff=True, is_superuser=False)
            .prefetch_related("groups")
            .order_by("username")
        )
        allowed_roles = ["ADAA", "STAFF"]

    else:
        users = (
            User.objects
            .filter(is_staff=True, groups__name="STAFF")
            .prefetch_related("groups")
            .order_by("username")
            .distinct()
        )
        allowed_roles = ["STAFF"]

    context = _admin_context(
        "user_management",
        {
            "users": users,
            "allowed_roles": allowed_roles,
            "current_admin_role": current_role,
        },
    )

    return render(request, "admin/admin_user_management.html", context)


def _expand_department_label(value):
    value = _clean_value(value)

    if not value:
        return ""

    return DEPARTMENT_MAP.get(value.upper(), value)

def _build_rank_lookup_from_records(rank_records):
    """
    Builds a searchable rank lookup from the uploaded rank Excel file.

    Expected rank file columns can be:
    Department | No. | Faculty | Rank | ...
    or:
    Name | Rank
    or:
    Faculty Name | Academic Rank
    """

    rank_lookup = {}

    for record in rank_records or []:
        name = _clean_value(record.get("name", ""))
        normalized_name = (
            record.get("normalized_name")
            or _normalize_person_name(name)
        )

        if not normalized_name:
            continue

        rank_value = _clean_value(record.get("rank", ""))
        department_value = _expand_department_label(record.get("department", ""))

        if not rank_value and not department_value:
            continue

        existing = rank_lookup.setdefault(normalized_name, {
            "name": name,
            "normalized_name": normalized_name,
            "rank": "",
            "department": "",
            "sheet": record.get("sheet", ""),
            "row": record.get("row", ""),
        })

        if rank_value and not existing["rank"]:
            existing["rank"] = rank_value

        if department_value and not existing["department"]:
            existing["department"] = department_value

    return rank_lookup

def _find_rank_data_for_person(possible_names, rank_lookup):
    """
    Finds rank data safely.

    IMPORTANT:
    This must always return a dictionary.
    If no rank match is found, return {} instead of None.
    """

    if not rank_lookup:
        return {}

    candidate_keys = list(rank_lookup.keys())
    normalized_candidates = []

    for value in possible_names or []:
        if not value:
            continue

        normalized = _normalize_person_name(value)

        if normalized and normalized not in normalized_candidates:
            normalized_candidates.append(normalized)

    # 1. Exact normalized match first
    for normalized in normalized_candidates:
        if normalized in rank_lookup:
            return rank_lookup.get(normalized, {})

    # 2. Safe fuzzy match
    used_keys = set()

    for normalized in normalized_candidates:
        matched_key, score = _find_fuzzy_match(
            normalized,
            candidate_keys,
            used_keys,
        )

        if matched_key and score >= 0.88:
            return rank_lookup.get(matched_key, {})

    # 3. Safer token containment fallback
    for normalized in normalized_candidates:
        source_tokens = _token_set(normalized)

        if not source_tokens:
            continue

        for candidate_key in candidate_keys:
            candidate_tokens = _token_set(candidate_key)

            if not candidate_tokens:
                continue

            common_tokens = source_tokens & candidate_tokens

            # Only allow fallback if first name and last name are safe.
            if len(common_tokens) >= 2 and _is_safe_name_match(normalized, candidate_key):
                return rank_lookup.get(candidate_key, {})

    # VERY IMPORTANT:
    # Without this, Python returns None and rank_data.get() will crash.
    return {}